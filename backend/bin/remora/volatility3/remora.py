#!/usr/bin/env python3
"""
Volatility3 GUI – Professional graphical frontend for Volatility3.

Usage:
    python3 remora.py

Requirements:
    pip install PyQt6
Optional:
    pip install openpyxl   (Excel export)
"""

import csv
import hashlib
import json
import os
import re as _re
import subprocess
import sys
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

SCRIPT_DIR = Path(__file__).parent.absolute()
sys.path.insert(0, str(SCRIPT_DIR))

# ---------------------------------------------------------------------------
# PyQt6
# ---------------------------------------------------------------------------
try:
    from PyQt6.QtCore import QProcess, QSettings, Qt, QThread, pyqtSignal
    from PyQt6.QtGui import QColor, QFont, QPageLayout, QPageSize, QTextCursor, QTextDocument
    from PyQt6.QtWidgets import (
        QAbstractItemView, QApplication, QCheckBox, QComboBox,
        QFileDialog, QFormLayout, QFrame, QGroupBox, QHBoxLayout,
        QHeaderView, QLabel, QLineEdit, QMainWindow, QMenu, QMessageBox,
        QProgressBar, QPushButton, QScrollArea, QSpinBox, QSplitter,
        QTabWidget, QTableWidget, QTableWidgetItem, QTextEdit,
        QTreeWidget, QTreeWidgetItem, QVBoxLayout, QWidget,
    )
    from PyQt6.QtGui import QAction
except ImportError:
    print("ERROR: PyQt6 is required.  pip install PyQt6")
    sys.exit(1)

try:
    from PyQt6.QtPrintSupport import QPrinter
    HAS_PRINT = True
except ImportError:
    HAS_PRINT = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font as XLFont, PatternFill
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

# ---------------------------------------------------------------------------
# Volatility3 discovery
# ---------------------------------------------------------------------------
VOL3_OK, VOL3_ERROR = False, ""
try:
    import volatility3.framework as vol_framework
    import volatility3.plugins as vol_plugins_module
    VOL3_OK = True
except Exception as exc:
    VOL3_ERROR = str(exc)

HIDDEN_REQ_TYPES = {
    "TranslationLayerRequirement", "SymbolTableRequirement",
    "ModuleRequirement", "VersionRequirement",
    "LayerListRequirement", "MultiRequirement", "PluginRequirement",
}
AUTO_FILLED_ARGS = {"single_location"}

# ---------------------------------------------------------------------------
# Triage profiles — named batteries of plugins for one-click triage. Entries
# are full plugin names; any not present in the loaded Volatility build are
# silently skipped, so a single profile can span Volatility versions.
# ---------------------------------------------------------------------------
TRIAGE_PROFILES: Dict[str, List[str]] = {
    "Windows · Quick Triage": [
        "windows.pslist", "windows.pstree", "windows.psscan",
        "windows.cmdline", "windows.netscan", "windows.netstat",
        "windows.malfind", "windows.dlllist", "windows.svcscan",
        "windows.registry.hashdump",
    ],
    "Windows · Ransomware": [
        "windows.pslist", "windows.cmdline", "windows.netscan",
        "windows.malfind", "windows.filescan", "windows.mutantscan",
        "windows.dlllist",
    ],
    "Windows · Rootkit / Hidden": [
        "windows.psscan", "windows.ssdt", "windows.modules",
        "windows.modscan", "windows.malfind", "windows.ldrmodules",
        "windows.callbacks", "windows.driverirp",
    ],
    "Linux · Quick Triage": [
        "linux.pslist", "linux.pstree", "linux.psaux", "linux.bash",
        "linux.lsof", "linux.malfind", "linux.check_syscall",
        "linux.check_modules",
    ],
    "macOS · Quick Triage": [
        "mac.pslist", "mac.pstree", "mac.lsof", "mac.bash",
        "mac.netstat", "mac.malfind",
    ],
}

# ===========================================================================
# MITRE ATT&CK mappings
# ===========================================================================

# Technique ID → human-readable name
MITRE_TECHNIQUES: Dict[str, str] = {
    # ── Credential Access ────────────────────────────────────────────────────
    "T1003":     "OS Credential Dumping",
    "T1003.001": "OS Credential Dumping: LSASS Memory",
    "T1003.002": "OS Credential Dumping: SAM",
    "T1003.004": "OS Credential Dumping: LSA Secrets",
    "T1003.005": "OS Credential Dumping: Cached Domain Credentials",
    "T1056.001": "Input Capture: Keylogging",
    "T1552.001": "Unsecured Credentials: Credentials In Files",
    "T1556.001": "Modify Auth Process: Domain Controller Authentication",
    # ── Defence Evasion ──────────────────────────────────────────────────────
    "T1014":     "Rootkit",
    "T1027":     "Obfuscated Files or Information",
    "T1027.007": "Obfuscated Files: Dynamic API Resolution",
    "T1036":     "Masquerading",
    "T1036.005": "Masquerading: Match Legitimate Name or Location",
    "T1070":     "Indicator Removal",
    "T1070.004": "Indicator Removal: File Deletion",
    "T1112":     "Modify Registry",
    "T1140":     "Deobfuscate/Decode Files or Information",
    "T1562.001": "Impair Defenses: Disable or Modify Tools",
    "T1562.006": "Impair Defenses: Indicator Blocking",
    "T1564.001": "Hide Artifacts: Hidden Files and Directories",
    "T1564.004": "Hide Artifacts: NTFS File Attributes (ADS)",
    "T1622":     "Debugger Evasion",
    # ── Discovery ────────────────────────────────────────────────────────────
    "T1007":     "System Service Discovery",
    "T1012":     "Query Registry",
    "T1016":     "System Network Configuration Discovery",
    "T1049":     "System Network Connections Discovery",
    "T1057":     "Process Discovery",
    "T1069":     "Permission Groups Discovery",
    "T1082":     "System Information Discovery",
    "T1083":     "File and Directory Discovery",
    "T1518":     "Software Discovery",
    # ── Execution ────────────────────────────────────────────────────────────
    "T1059":     "Command and Scripting Interpreter",
    "T1059.001": "Command and Scripting Interpreter: PowerShell",
    "T1059.003": "Command and Scripting Interpreter: Windows Command Shell",
    "T1059.004": "Command and Scripting Interpreter: Unix Shell",
    "T1106":     "Native API",
    "T1129":     "Shared Modules",
    # ── Hijack Execution Flow ────────────────────────────────────────────────
    "T1574":     "Hijack Execution Flow",
    "T1574.001": "Hijack Execution Flow: DLL Search Order Hijacking",
    "T1574.002": "Hijack Execution Flow: DLL Side-Loading",
    "T1574.006": "Hijack Execution Flow: Dynamic Linker Hijacking",
    # ── Impact ───────────────────────────────────────────────────────────────
    "T1486":     "Data Encrypted for Impact",
    # ── Collection ───────────────────────────────────────────────────────────
    "T1005":     "Data from Local System",
    # ── Lateral Movement ─────────────────────────────────────────────────────
    "T1021":     "Remote Services",
    "T1563":     "Remote Service Session Hijacking",
    # ── Persistence ──────────────────────────────────────────────────────────
    "T1053.005": "Scheduled Task/Job: Scheduled Task",
    "T1543.003": "Create or Modify System Process: Windows Service",
    "T1546.011": "Event Triggered Execution: Application Shimming",
    "T1547.001": "Boot or Logon Autostart: Registry Run Keys / Startup Folder",
    "T1542.003": "Pre-OS Boot: Bootkit",
    # ── Privilege Escalation ─────────────────────────────────────────────────
    "T1055":     "Process Injection",
    "T1055.001": "Process Injection: DLL Injection",
    "T1055.002": "Process Injection: Portable Executable Injection",
    "T1055.003": "Process Injection: Thread Execution Hijacking",
    "T1055.012": "Process Injection: Process Hollowing",
    "T1055.015": "Process Injection: ListPlanting / Process Ghosting",
    "T1068":     "Exploitation for Privilege Escalation",
    "T1134":     "Access Token Manipulation",
    "T1620":     "Reflective Code Loading",
    # ── Command and Control ──────────────────────────────────────────────────
    "T1071":     "Application Layer Protocol (C2)",
    "T1105":     "Ingress Tool Transfer",
    # ── Credential Access / Subvert Trust ────────────────────────────────────
    "T1218":     "System Binary Proxy Execution",
    "T1553":     "Subvert Trust Controls",
    "T1553.004": "Subvert Trust Controls: Install Root Certificate",
    # ── Misc ─────────────────────────────────────────────────────────────────
    "T1078":     "Valid Accounts",
}

# Plugin name segment(s) → list of MITRE technique IDs.
# Keys are matched against the dot-separated segments of a plugin's full name
# (e.g. "hashdump" matches both "windows.hashdump" and "windows.registry.hashdump").
# These encode what adversary behaviour the plugin *detects evidence of*, NOT what
# the plugin is doing itself.  Plugins that are pure forensic infrastructure
# (info, crashinfo, statistics, virtmap, poolscanner, etc.) are intentionally
# absent — they have no meaningful ATT&CK adversary-technique mapping.
PLUGIN_MITRE_MAP: Dict[str, List[str]] = {
    # ── Process discovery ───────────────────────────────────────────────────
    "pslist":               ["T1057", "T1036.005"],
    "psscan":               ["T1057", "T1014", "T1036"],
    "pstree":               ["T1057", "T1059.001", "T1059.003", "T1036.005"],
    "psxview":              ["T1057", "T1014"],
    "psaux":                ["T1057"],
    "pidhashtable":         ["T1057"],
    "proc":                 ["T1057"],
    "pscallstack":          ["T1057", "T1055"],
    # ── Process injection / code injection / hollowing ──────────────────────
    "malfind":              ["T1055", "T1055.001", "T1055.002", "T1055.012",
                             "T1620", "T1027.007"],
    "hollowprocesses":      ["T1055.012", "T1055"],
    "processghosting":      ["T1055.015", "T1055", "T1014"],
    "ptrace":               ["T1055"],
    "pebmasquerade":        ["T1036.005", "T1055"],
    "suspicious_threads":   ["T1055", "T1055.003"],
    "orphan_kernel_threads":["T1055", "T1014"],
    "debugregisters":       ["T1622", "T1055"],
    "vmaregexscan":         ["T1055", "T1027"],
    "vadinfo":              ["T1055", "T1620", "T1140"],
    "vadwalk":              ["T1055", "T1620"],
    "vadregexscan":         ["T1055", "T1027"],
    "vadyarascan":          ["T1055", "T1027", "T1059"],
    "vmayarascan":          ["T1055", "T1027", "T1005"],
    "threads":              ["T1055", "T1055.003"],
    "thrdscan":             ["T1055", "T1055.003", "T1055.004"],
    "suspended_threads":    ["T1055"],
    "proc_maps":            ["T1055", "T1620"],
    "elfs":                 ["T1083"],
    "memmap":               ["T1055", "T1620", "T1005"],
    "pedump":               ["T1055", "T1620", "T1005"],
    "pe_symbols":           ["T1027.007", "T1055"],
    # ── Credential dumping ──────────────────────────────────────────────────
    "hashdump":             ["T1003", "T1003.002"],
    "cachedump":            ["T1003", "T1003.005"],
    "lsadump":              ["T1003.004"],
    "check_creds":          ["T1003"],
    "skeleton_key_check":   ["T1556.001"],
    "truecrypt":            ["T1552.001", "T1027", "T1486"],
    # ── Registry ────────────────────────────────────────────────────────────
    "hivelist":             ["T1012"],
    "hivescan":             ["T1012", "T1112"],
    "printkey":             ["T1012", "T1547.001", "T1112"],
    "userassist":           ["T1012", "T1059"],
    "getcellroutine":       ["T1014", "T1562.001"],
    "certificates":         ["T1553.004"],
    "amcache":              ["T1059", "T1218", "T1036"],
    "shimcachemem":         ["T1546.011", "T1059", "T1218"],
    "scheduled_tasks":      ["T1053.005"],
    # ── Command / shell ─────────────────────────────────────────────────────
    "cmdline":              ["T1059", "T1059.001", "T1059.003", "T1218", "T1027"],
    "cmdscan":              ["T1059", "T1059.003", "T1562"],
    "consoles":             ["T1059", "T1059.003", "T1105", "T1057"],
    "bash":                 ["T1059.004"],
    "joblinks":             ["T1059"],
    "kthreads":             ["T1059"],
    # ── Network / lateral movement ──────────────────────────────────────────
    "netscan":              ["T1049", "T1071", "T1021"],
    "netstat":              ["T1049", "T1071", "T1021"],
    "sockstat":             ["T1049"],
    "sockscan":             ["T1049"],
    "ip":                   ["T1016"],
    "ifconfig":             ["T1016"],
    "handles":              ["T1083", "T1016", "T1057", "T1012", "T1082"],
    "sessions":             ["T1563", "T1078"],
    # ── Rootkit / defence evasion ────────────────────────────────────────────
    "ssdt":                 ["T1014", "T1562.001"],
    "callbacks":            ["T1014"],
    "check_afinfo":         ["T1014"],
    "check_idt":            ["T1014"],
    "check_modules":        ["T1014"],
    "check_syscall":        ["T1014"],
    "check_sysctl":         ["T1014"],
    "check_trap_table":     ["T1014"],
    "hidden_modules":       ["T1014"],
    "modxview":             ["T1014"],
    "modscan":              ["T1014"],
    "modules":              ["T1014", "T1543.003"],
    "module_extract":       ["T1014"],
    "lsmod":                ["T1082", "T1014"],
    "timers":               ["T1014", "T1543.003"],
    "kauth_listeners":      ["T1014"],
    "kauth_scopes":         ["T1014"],
    "socket_filters":       ["T1014"],
    "netfilter":            ["T1014"],
    "trustedbsd":           ["T1553"],
    "ebpf":                 ["T1014", "T1055"],
    "devicetree":           ["T1014"],
    "driverirp":            ["T1014", "T1543.003"],
    "driverscan":           ["T1014", "T1543.003"],
    "drivermodule":         ["T1014"],
    "unloadedmodules":      ["T1014", "T1070"],
    "unhooked_system_calls":["T1106", "T1055", "T1562.001"],
    "direct_system_calls":  ["T1106", "T1055", "T1562.001"],
    "indirect_system_calls":["T1106", "T1055", "T1562.001"],
    "etwpatch":             ["T1562.006", "T1562.001"],
    "ftrace":               ["T1014", "T1056.001"],
    "perf_events":          ["T1014"],
    "tracepoints":          ["T1014"],
    "tracing":              ["T1014"],
    # ── Keylogging ──────────────────────────────────────────────────────────
    "keyboard_notifiers":   ["T1056.001"],
    "tty_check":            ["T1056.001", "T1014"],
    # ── Privilege / token ───────────────────────────────────────────────────
    "privileges":           ["T1134"],
    "getsids":              ["T1069", "T1134"],
    "getservicesids":       ["T1069"],
    "capabilities":         ["T1134", "T1068"],
    # ── Persistence / services ──────────────────────────────────────────────
    "svcscan":              ["T1007", "T1543.003"],
    "svclist":              ["T1007", "T1543.003"],
    "svcdiff":              ["T1543.003"],
    # ── DLL / module hijacking ──────────────────────────────────────────────
    "dlllist":              ["T1055.001", "T1055", "T1574", "T1129"],
    "ldrmodules":           ["T1055.001", "T1055", "T1574.001"],
    "iat":                  ["T1574", "T1055", "T1027.007"],
    "library_list":         ["T1574.006"],
    # ── File / MFT / data collection ────────────────────────────────────────
    "filescan":             ["T1083", "T1005"],
    "mftscan":              ["T1083", "T1070.004", "T1564.001", "T1564.004"],
    "dumpfiles":            ["T1005"],
    "lsof":                 ["T1083", "T1049"],
    "list_files":           ["T1083"],
    "vfsevents":            ["T1083"],
    "strings":              ["T1027", "T1059", "T1071"],
    "pagecache":            ["T1005"],
    "regexscan":            ["T1005"],
    "symlinkscan":          ["T1083", "T1564.001"],
    "mutantscan":           ["T1071", "T1105"],
    "mountinfo":            ["T1082", "T1083"],
    "mount":                ["T1082"],
    # ── System information discovery (adversary-reachable data) ─────────────
    "envars":               ["T1082", "T1059"],
    "iomem":                ["T1082"],
    "kallsyms":             ["T1082"],
    "kmsg":                 ["T1082"],
    "boottime":             ["T1082"],
    "vmcoreinfo":           ["T1082"],
    "desktops":             ["T1082"],
    "deskscan":             ["T1082"],
    "windowstations":       ["T1082"],
    "dmesg":                ["T1082"],
    "kevents":              ["T1082"],
    "fbdev":                ["T1082"],
    "bigpools":             ["T1082", "T1014"],
    "timeliner":            ["T1082"],
    # ── Masquerading ────────────────────────────────────────────────────────
    "verinfo":              ["T1036"],
    "process_spoofing":     ["T1036", "T1036.005"],
    # ── Bootkit ─────────────────────────────────────────────────────────────
    "mbrscan":              ["T1542.003"],
    # ── Discovery / scanning ────────────────────────────────────────────────
    "yarascan":             ["T1518", "T1027"],
    # ── No ATT&CK mapping (forensic infrastructure only) ────────────────────
    # info, crashinfo, statistics, virtmap, poolscanner, kpcrs,
    # pe_symbols (infrastructure), configwriter, layerwriter,
    # isfinfo, frameworkinfo, banners, vmscan — intentionally omitted.
}

# Threat actor/group → MITRE technique IDs they are known to use in the wild.
# Sources: MITRE ATT&CK Groups, published threat reports.
THREAT_ACTORS: Dict[str, List[str]] = {
    "APT1 (Comment Crew / Unit 61398)":
        ["T1059", "T1082", "T1057", "T1083", "T1049", "T1012"],
    "APT28 (Fancy Bear / Sofacy / Pawn Storm)":
        ["T1055", "T1059", "T1059.001", "T1082", "T1003", "T1014",
         "T1547.001", "T1574", "T1027"],
    "APT29 (Cozy Bear / The Dukes)":
        ["T1055.012", "T1059", "T1059.001", "T1082", "T1003", "T1003.001",
         "T1547.001", "T1027", "T1620"],
    "APT32 (OceanLotus / Cobalt Kitty)":
        ["T1055", "T1059", "T1082", "T1574.001", "T1027"],
    "APT38 / Lazarus Group (Hidden Cobra)":
        ["T1055", "T1059", "T1082", "T1003", "T1014", "T1486", "T1036",
         "T1620"],
    "APT41 (Winnti / BARIUM / Double Dragon)":
        ["T1055", "T1059", "T1003", "T1082", "T1014", "T1053.005",
         "T1574", "T1106"],
    "BlackCat / ALPHV":
        ["T1486", "T1082", "T1083", "T1003", "T1003.001"],
    "Carbanak / FIN7 / Navigator Group":
        ["T1059", "T1059.001", "T1055", "T1082", "T1003", "T1543.003",
         "T1547.001"],
    "Cl0p Ransomware":
        ["T1486", "T1082", "T1083", "T1003", "T1036"],
    "Conti Ransomware":
        ["T1486", "T1082", "T1083", "T1003", "T1003.001", "T1059",
         "T1059.001", "T1543.003"],
    "DarkHotel (Tapaoux)":
        ["T1059", "T1055", "T1082", "T1574.001"],
    "Equation Group (NSA / GCHQ-linked)":
        ["T1014", "T1082", "T1003", "T1055", "T1542.003", "T1027",
         "T1106"],
    "Gamaredon (Primitive Bear)":
        ["T1059", "T1082", "T1083", "T1547.001"],
    "Hive Ransomware":
        ["T1486", "T1082", "T1083", "T1003", "T1059"],
    "Kimsuky (Thallium / Black Banshee)":
        ["T1059", "T1082", "T1083", "T1056.001", "T1003"],
    "LockBit Ransomware":
        ["T1486", "T1082", "T1083", "T1003", "T1059", "T1543.003",
         "T1070"],
    "MuddyWater (Static Kitten)":
        ["T1059", "T1059.001", "T1082", "T1083", "T1027", "T1055"],
    "NotPetya / Sandworm (GRU Unit 74455)":
        ["T1486", "T1003", "T1003.001", "T1082", "T1059", "T1106"],
    "REvil / Sodinokibi":
        ["T1486", "T1082", "T1083", "T1003", "T1059", "T1547.001"],
    "Ryuk Ransomware":
        ["T1486", "T1082", "T1083", "T1003", "T1003.001", "T1059",
         "T1543.003", "T1070"],
    "ShadowPad (APT41-linked)":
        ["T1055", "T1082", "T1014", "T1574", "T1106"],
    "TA505 (Evil Corp-linked)":
        ["T1059", "T1059.001", "T1055", "T1082", "T1543.003", "T1027"],
    "Turla (Venomous Bear / Waterbug)":
        ["T1055", "T1059", "T1014", "T1082", "T1056.001", "T1574",
         "T1106"],
    "WannaCry (Lazarus Group)":
        ["T1486", "T1082", "T1059"],
    "Winnti Group (APT41 overlap)":
        ["T1055", "T1014", "T1082", "T1574", "T1543.003"],
    "Wizard Spider (Ryuk / TrickBot)":
        ["T1059", "T1059.001", "T1082", "T1003", "T1003.001", "T1486",
         "T1543.003", "T1070"],
}


# ATT&CK tactic groupings (tactic label → technique ID prefixes that belong to it)
MITRE_TACTICS: Dict[str, List[str]] = {
    "Execution":      ["T1059", "T1106", "T1129"],
    "Persistence":    ["T1053.005", "T1542.003", "T1543.003", "T1546.011",
                       "T1547.001"],
    "Priv Escalation":["T1055", "T1068", "T1134", "T1620"],
    "Def Evasion":    ["T1014", "T1027", "T1036", "T1070", "T1112", "T1140",
                       "T1562", "T1564", "T1622"],
    "Cred Access":    ["T1003", "T1056.001", "T1552.001", "T1556.001"],
    "Discovery":      ["T1007", "T1012", "T1016", "T1049", "T1057", "T1069",
                       "T1082", "T1083", "T1518"],
    "Lat Movement":   ["T1021", "T1563"],
    "Collection":     ["T1005"],
    "C2":             ["T1071", "T1105"],
    "Impact":         ["T1486"],
}

# Plugin name key → technique → confidence level ('H' high / 'M' medium / 'L' low).
# High  = plugin was specifically designed for this detection
# Medium = strong secondary signal (default when not listed)
# Low   = circumstantial / indirect evidence only
PLUGIN_MITRE_CONFIDENCE: Dict[str, Dict[str, str]] = {
    "malfind":              {"T1055": "H", "T1055.001": "H", "T1055.002": "H",
                             "T1055.012": "H", "T1620": "H", "T1027.007": "M"},
    "hollowprocesses":      {"T1055.012": "H"},
    "processghosting":      {"T1055.015": "H"},
    "skeleton_key_check":   {"T1556.001": "H"},
    "hashdump":             {"T1003.002": "H", "T1003": "H"},
    "cachedump":            {"T1003.005": "H", "T1003": "H"},
    "lsadump":              {"T1003.004": "H"},
    "check_creds":          {"T1003": "H"},
    "keyboard_notifiers":   {"T1056.001": "H"},
    "tty_check":            {"T1056.001": "H", "T1014": "H"},
    "ssdt":                 {"T1014": "H", "T1562.001": "H"},
    "callbacks":            {"T1014": "H"},
    "direct_system_calls":  {"T1106": "H", "T1562.001": "H"},
    "indirect_system_calls":{"T1106": "H", "T1562.001": "H"},
    "unhooked_system_calls":{"T1106": "H", "T1562.001": "H"},
    "etwpatch":             {"T1562.006": "H", "T1562.001": "H"},
    "debugregisters":       {"T1622": "H"},
    "mbrscan":              {"T1542.003": "H"},
    "pebmasquerade":        {"T1036.005": "H"},
    "process_spoofing":     {"T1036": "H", "T1036.005": "H"},
    "getcellroutine":       {"T1014": "H"},
    "trustedbsd":           {"T1553": "H"},
    "ebpf":                 {"T1014": "H"},
    "check_afinfo":         {"T1014": "H"},
    "check_idt":            {"T1014": "H"},
    "check_modules":        {"T1014": "H"},
    "check_syscall":        {"T1014": "H"},
    "check_sysctl":         {"T1014": "H"},
    "check_trap_table":     {"T1014": "H"},
    "hidden_modules":       {"T1014": "H"},
    "modxview":             {"T1014": "H"},
    "netfilter":            {"T1014": "H"},
    "ftrace":               {"T1014": "H", "T1056.001": "M"},
    "psxview":              {"T1014": "H", "T1057": "H"},
    "drivermodule":         {"T1014": "H"},
    "ldrmodules":           {"T1055.001": "H", "T1574.001": "H"},
    "shimcachemem":         {"T1546.011": "H"},
    "scheduled_tasks":      {"T1053.005": "H"},
    "svcdiff":              {"T1543.003": "H"},
    "mftscan":              {"T1070.004": "H", "T1564.004": "H", "T1564.001": "M"},
    "unloadedmodules":      {"T1070": "H"},
    "mutantscan":           {"T1071": "M", "T1105": "L"},
    "netscan":              {"T1049": "H", "T1071": "M", "T1021": "M"},
    "netstat":              {"T1049": "H", "T1071": "M", "T1021": "M"},
    "sessions":             {"T1563": "M", "T1078": "M"},
    "dlllist":              {"T1055.001": "H", "T1574": "M", "T1129": "M"},
    "vadinfo":              {"T1055": "M", "T1620": "M", "T1140": "L"},
    "cmdline":              {"T1059": "H", "T1059.001": "H", "T1059.003": "H",
                             "T1027": "L", "T1218": "M"},
    "cmdscan":              {"T1059.003": "H"},
    "consoles":             {"T1059.003": "H", "T1105": "L", "T1057": "L"},
    "certificates":         {"T1553.004": "H"},
    "truecrypt":            {"T1552.001": "H", "T1486": "M"},
    "symlinkscan":          {"T1564.001": "M"},
    "strings":              {"T1027": "M", "T1059": "L", "T1071": "L"},
    "yarascan":             {"T1518": "M"},
}


def _load_mitre_overrides() -> Optional[str]:
    """Merge an optional ``remora_mitre.json`` sitting next to this script into
    the inline MITRE maps so analysts can extend technique names, plugin→ATT&CK
    mappings, threat actors, tactics, and confidences without editing code.

    The file is a JSON object with any of these top-level keys, each a mapping
    that is merged into (and overrides) the corresponding inline default::

        {
          "techniques":  {"T1234": "Some Technique"},
          "plugin_map":  {"mything": ["T1234"]},
          "threat_actors": {"FIN-X": ["T1003", "T1055"]},
          "tactics":     {"TA0001": ["T1234"]},
          "confidence":  {"mything": {"T1234": "H"}}
        }

    Returns a short status string for logging, or ``None`` if no file present.
    """
    path = SCRIPT_DIR / "remora_mitre.json"
    if not path.is_file():
        return None
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:
        return f"remora_mitre.json ignored (parse error: {exc})"

    counts = []
    for key, target in (
        ("techniques", MITRE_TECHNIQUES),
        ("plugin_map", PLUGIN_MITRE_MAP),
        ("threat_actors", THREAT_ACTORS),
        ("tactics", MITRE_TACTICS),
        ("confidence", PLUGIN_MITRE_CONFIDENCE),
    ):
        section = data.get(key)
        if isinstance(section, dict):
            target.update(section)
            counts.append(f"{len(section)} {key}")
    return "remora_mitre.json loaded: " + ", ".join(counts) if counts else None


# Applied once at import; logged to the UI on startup so the merge is auditable.
MITRE_OVERRIDE_STATUS: Optional[str] = _load_mitre_overrides()


def plugin_in_profile(plugin_name: str, entries) -> bool:
    """Return True if a discovered plugin matches any triage-profile entry.

    Volatility plugin names carry a trailing class segment
    (e.g. ``windows.pslist.PsList``); profile entries omit it
    (``windows.pslist``). Matching therefore accepts an exact name or a
    dotted-prefix, while the trailing ``.`` guard prevents
    ``windows.pslist`` from spuriously matching ``windows.pslister``.
    """
    nl = plugin_name.lower()
    return any(nl == e.lower() or nl.startswith(e.lower() + ".") for e in entries)


def _get_confidence(plugin_key: str, technique_id: str) -> str:
    """Return H/M/L confidence for a plugin-key + technique pair."""
    return PLUGIN_MITRE_CONFIDENCE.get(plugin_key, {}).get(technique_id, "M")


def _get_plugin_techniques(plugin_name: str) -> List[str]:
    """Return all MITRE technique IDs that a plugin maps to.

    Matching is done by checking whether any dot-separated segment of the
    *full* plugin name (e.g. ``windows.registry.hashdump``) equals a key in
    :data:`PLUGIN_MITRE_MAP`.  Multi-segment keys (none currently) would
    require consecutive segments to match.
    """
    parts = plugin_name.lower().split(".")
    techniques: List[str] = []
    seen: set = set()
    for key, techs in PLUGIN_MITRE_MAP.items():
        key_parts = key.split(".")
        klen = len(key_parts)
        # Slide a window of size klen over parts
        for i in range(len(parts) - klen + 1):
            if parts[i:i + klen] == key_parts:
                for t in techs:
                    if t not in seen:
                        techniques.append(t)
                        seen.add(t)
                break
    return techniques


# ===========================================================================
# Theme engine
# ===========================================================================

# ── Palettes ─────────────────────────────────────────────────────────────────
#
# Dark: modelled after JetBrains Darcula / VS Code One Dark Pro.
#       NOT pitch-black – everything is a visible grey-blue.
# Light: clean neutral white-grey.

_DARK = dict(
    # ── backgrounds ──────────────────────────────────────────────────────
    bg          = "#0d0f14",   # window / content area
    bg_panel    = "#11141a",   # side panels, plugin browser
    bg_elev     = "#181c24",   # inputs, elevated surfaces
    bg_hover    = "#1e2430",   # hover highlight
    bg_sel      = "#0c2a48",   # selected row / tree item
    bg_header   = "#090b10",   # header bar, tab strip, status bar
    bg_alt      = "#101318",   # alternate table row
    bg_code     = "#080a0d",   # log / monospace area
    # ── borders ──────────────────────────────────────────────────────────
    border      = "#222840",   # visible borders
    border_dim  = "#161a28",   # subtle dividers
    # ── accent ───────────────────────────────────────────────────────────
    accent      = "#1f9cd8",
    accent_hov  = "#1880b4",
    accent_prs  = "#106490",
    # ── text  (all clearly readable on dark backgrounds) ─────────────────
    text_hi     = "#dce8f8",   # headings, active labels
    text_body   = "#96aec8",   # primary readable text
    text_sec    = "#546880",   # secondary labels
    text_muted  = "#324050",   # muted / placeholders
    text_dim    = "#202c38",   # barely-there captions
    # ── semantic ─────────────────────────────────────────────────────────
    success     = "#2eb86a",
    warning     = "#c8880e",
    error       = "#c83040",
    # ── misc ─────────────────────────────────────────────────────────────
    sb_handle   = "#222a38",
    sb_hover    = "#2e3a4e",
    drop_bg     = "#0d0f14",
    drop_border = "#222840",
)

_LIGHT = dict(
    bg          = "#f0f3f7",
    bg_panel    = "#e4eaf2",
    bg_elev     = "#ffffff",
    bg_hover    = "#d8e2ee",
    bg_sel      = "#b0cce6",
    bg_header   = "#dce4ef",
    bg_alt      = "#eaeff6",
    bg_code     = "#e2e8f2",
    border      = "#94a8be",
    border_dim  = "#bccad8",
    accent      = "#1470a8",
    accent_hov  = "#0e5888",
    accent_prs  = "#0a4068",
    text_hi     = "#060c14",
    text_body   = "#162030",
    text_sec    = "#304858",
    text_muted  = "#5a7080",
    text_dim    = "#8898a8",
    success     = "#0e7840",
    warning     = "#7a4a00",
    error       = "#962030",
    sb_handle   = "#94a8be",
    sb_hover    = "#7088a0",
    drop_bg     = "#e4eaf2",
    drop_border = "#94a8be",
)

# Mutable reference – updated by _toggle_theme()
_ACTIVE: dict = dict(_DARK)

def _c() -> dict:
    """Return the active colour palette."""
    return _ACTIVE

# ── CSS template ──────────────────────────────────────────────────────────────
# ${key} placeholders; no conflict with CSS syntax.

_CSS = r"""
* {
    font-family: "Inter", "SF Pro Text", "Segoe UI", Ubuntu, sans-serif;
    font-size: 12px;
    outline: 0;
}
QMainWindow, QDialog, QWidget {
    background-color: ${bg};
    color: ${text_body};
}

/* ── Menu ──────────────────────────────────────────────────────────────── */
QMenuBar {
    background-color: ${bg_header};
    color: ${text_sec};
    border-bottom: 1px solid ${border_dim};
    padding: 0;
    min-height: 24px;
}
QMenuBar::item { padding: 4px 12px; background: transparent; }
QMenuBar::item:selected { background-color: ${bg_hover}; color: ${text_hi}; }
QMenu {
    background-color: ${bg_panel};
    color: ${text_body};
    border: 1px solid ${border};
    padding: 2px 0;
}
QMenu::item { padding: 5px 24px 5px 14px; }
QMenu::item:selected { background-color: ${bg_hover}; color: ${text_hi}; }
QMenu::item:disabled { color: ${text_dim}; }
QMenu::separator { height: 1px; background: ${border_dim}; margin: 3px 0; }

/* ── Splitters ─────────────────────────────────────────────────────────── */
QSplitter::handle:horizontal { width: 1px; background: ${border_dim}; }
QSplitter::handle:vertical   { height: 1px; background: ${border_dim}; }

/* ── Tree widget ───────────────────────────────────────────────────────── */
QTreeWidget {
    background-color: ${bg_panel};
    color: ${text_sec};
    border: none;
    show-decoration-selected: 1;
    font-size: 13px;
}
QTreeWidget::item { padding: 4px 6px; border: none; min-height: 24px; }
QTreeWidget::item:hover   { background-color: ${bg_hover}; color: ${text_body}; }
QTreeWidget::item:selected {
    background-color: ${bg_sel};
    color: ${text_hi};
    border-left: 2px solid ${accent};
    padding-left: 4px;
}
QTreeWidget::branch { background: transparent; }
QTreeWidget QHeaderView::section {
    background-color: ${bg_header};
    color: ${text_muted};
    padding: 4px 8px;
    border: none;
    border-bottom: 1px solid ${border_dim};
    font-size: 10px; font-weight: 700; letter-spacing: 1px;
    text-transform: uppercase;
}

/* ── Tabs ──────────────────────────────────────────────────────────────── */
QTabWidget::pane { border: none; border-top: 1px solid ${border_dim}; background: ${bg}; }
QTabBar { background: ${bg_header}; }
QTabBar::tab {
    background: transparent;
    color: ${text_muted};
    padding: 6px 16px 5px;
    border: none;
    border-right: 1px solid ${border_dim};
    border-top: 2px solid transparent;
    min-width: 70px;
    font-size: 11px;
    letter-spacing: 0.3px;
}
QTabBar::tab:first { border-left: 1px solid ${border_dim}; }
QTabBar::tab:selected {
    color: ${text_hi};
    background: ${bg};
    border-top: 2px solid ${accent};
}
QTabBar::tab:hover:!selected { color: ${text_sec}; background: ${bg_hover}; }
QTabBar::close-button { subcontrol-position: right; margin: 0 2px; }

/* ── Table ─────────────────────────────────────────────────────────────── */
QTableWidget {
    background-color: ${bg};
    color: ${text_body};
    border: none;
    gridline-color: ${border_dim};
    selection-background-color: ${bg_sel};
    alternate-background-color: ${bg_alt};
    font-family: "JetBrains Mono", "Fira Code", "Cascadia Code", "Consolas", monospace;
    font-size: 11px;
}
QTableWidget::item { padding: 1px 10px; border: none; min-height: 22px; }
QTableWidget::item:selected { background-color: ${bg_sel}; color: ${text_hi}; }
QHeaderView::section {
    background-color: ${bg_header};
    color: ${text_muted};
    padding: 4px 10px;
    border: none;
    border-bottom: 1px solid ${border};
    border-right: 1px solid ${border_dim};
    font-size: 10px; font-weight: 700; letter-spacing: 1px;
    text-transform: uppercase; min-height: 24px;
}
QHeaderView::section:last-child { border-right: none; }
QHeaderView::section:hover { color: ${text_body}; }

/* ── Primary button ────────────────────────────────────────────────────── */
QPushButton {
    background-color: ${accent};
    color: #ffffff;
    border: none;
    border-radius: 2px;
    padding: 5px 16px;
    font-weight: 600;
    font-size: 11px;
    letter-spacing: 0.4px;
    min-height: 26px;
}
QPushButton:hover   { background-color: ${accent_hov}; }
QPushButton:pressed { background-color: ${accent_prs}; }
QPushButton:disabled { background-color: ${bg_elev}; color: ${text_dim}; }

/* ── Ghost button ──────────────────────────────────────────────────────── */
QPushButton[class="ghost"] {
    background-color: transparent;
    color: ${text_sec};
    border: 1px solid ${border};
    border-radius: 2px;
}
QPushButton[class="ghost"]:hover { background-color: ${bg_hover}; color: ${text_body}; }
QPushButton[class="ghost"]:pressed { background-color: ${bg_elev}; }
QPushButton[class="ghost"]:disabled { color: ${text_dim}; border-color: ${border_dim}; }

/* ── Inputs ────────────────────────────────────────────────────────────── */
QLineEdit, QSpinBox, QComboBox {
    background-color: ${bg_elev};
    color: ${text_body};
    border: 1px solid ${border};
    border-radius: 2px;
    padding: 4px 8px;
    font-size: 12px;
    min-height: 24px;
    selection-background-color: ${bg_sel};
}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus { border-color: ${accent}; }
QLineEdit::placeholder { color: ${text_muted}; }
QSpinBox::up-button, QSpinBox::down-button {
    background-color: ${bg_hover}; border: none; width: 16px;
}
QComboBox::drop-down { border: none; width: 20px; background: ${bg_hover}; }
QComboBox QAbstractItemView {
    background-color: ${bg_elev}; color: ${text_body};
    border: 1px solid ${border}; selection-background-color: ${bg_sel};
}

/* ── Checkbox ──────────────────────────────────────────────────────────── */
QCheckBox { color: ${text_body}; spacing: 7px; }
QCheckBox::indicator {
    width: 13px; height: 13px; border-radius: 1px;
    border: 1px solid ${border}; background-color: ${bg_elev};
}
QCheckBox::indicator:checked { background-color: ${accent}; border-color: ${accent}; }
QCheckBox::indicator:hover   { border-color: ${accent}; }

/* ── Scrollbars ────────────────────────────────────────────────────────── */
QScrollBar:vertical   { background: ${bg}; width: 6px;  border: none; margin: 0; }
QScrollBar:horizontal { background: ${bg}; height: 6px; border: none; margin: 0; }
QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
    background: ${sb_handle}; border-radius: 3px; min-height: 20px; min-width: 20px; margin: 1px;
}
QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover { background: ${sb_hover}; }
QScrollBar::add-line, QScrollBar::sub-line { height: 0; width: 0; }

/* ── Log / code area ───────────────────────────────────────────────────── */
QTextEdit, QPlainTextEdit {
    background-color: ${bg_code};
    color: ${text_sec};
    border: none;
    font-family: "JetBrains Mono", "Fira Code", "Cascadia Code", "Consolas", monospace;
    font-size: 11px;
}

/* ── Progress bar ──────────────────────────────────────────────────────── */
QProgressBar { background-color: ${bg_elev}; border: none; height: 2px; border-radius: 0; }
QProgressBar::chunk { background-color: ${accent}; border-radius: 0; }

/* ── Status bar ────────────────────────────────────────────────────────── */
QStatusBar {
    background-color: ${bg_header}; color: ${text_muted};
    border-top: 1px solid ${border_dim}; font-size: 11px; padding: 0 10px;
}
QStatusBar::item { border: none; }

/* ── Group boxes ───────────────────────────────────────────────────────── */
QGroupBox {
    border: 1px solid ${border_dim}; margin-top: 18px; padding-top: 8px;
}
QGroupBox::title {
    subcontrol-origin: margin; left: 10px; padding: 0 5px;
    color: ${text_muted}; background: ${bg};
    font-size: 10px; font-weight: 700; letter-spacing: 1.2px; text-transform: uppercase;
}

/* ── Tooltip ───────────────────────────────────────────────────────────── */
QToolTip {
    background-color: ${bg_elev}; color: ${text_body};
    border: 1px solid ${accent}; border-radius: 1px; padding: 4px 9px; font-size: 11px;
}

/* ── Form labels ───────────────────────────────────────────────────────── */
QFormLayout > QLabel { color: ${text_sec}; font-size: 11px; padding-right: 4px; }

/* ── Scroll area ───────────────────────────────────────────────────────── */
QScrollArea { border: none; background: transparent; }
QScrollArea > QWidget > QWidget { background: transparent; }

/* ── Named / special widgets ───────────────────────────────────────────── */
QFrame#dropZone {
    border: 1px solid ${border};
    border-left: 3px solid ${accent};
    background-color: ${drop_bg};
}
QWidget#sidePanel {
    background-color: ${bg_panel};
    border-right: 1px solid ${border_dim};
}
QWidget#headerBar {
    background-color: ${bg_header};
    border-bottom: 1px solid ${border};
}
QWidget#subHeader {
    background-color: ${bg_header};
    border-bottom: 1px solid ${border_dim};
}
QWidget#tabToolbar {
    background-color: ${bg_header};
    border-bottom: 1px solid ${border_dim};
}
QWidget#tabFooter {
    background-color: ${bg_header};
    border-top: 1px solid ${border_dim};
}
QWidget#pluginInfo {
    background-color: ${bg_panel};
    border-bottom: 1px solid ${border_dim};
}
QWidget#configFooter {
    background-color: ${bg_header};
    border-top: 1px solid ${border_dim};
}
QLabel#sectionLabel {
    color: ${accent};
    font-size: 10px; font-weight: 700; letter-spacing: 1.5px; text-transform: uppercase;
}
QLabel#pluginTitle {
    color: ${text_hi};
    font-size: 13px; font-weight: 700;
}
QLabel#pluginDesc  { color: ${text_sec}; font-size: 11px; }
QLabel#fileChip    { color: ${text_muted}; font-size: 11px; }
QLabel#fileLoaded  { color: ${success}; font-size: 11px; font-weight: 600; }
QLabel#runLabel    { color: ${warning}; font-size: 11px; }
QLabel#tabInfo     { color: ${text_sec}; font-size: 11px; }
QLabel#tabCount    { color: ${text_dim}; font-size: 10px; }
QLabel#tabTs       { color: ${border_dim}; font-size: 10px; }
QLabel#dropMain    { color: ${text_sec}; font-size: 11px; }
QLabel#dropFile    { color: ${success}; font-size: 11px; font-weight: 600; }
QLabel#catLabel    { color: ${accent}; font-size: 10px; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; }
QLabel#noneLabel   { color: ${text_dim}; font-style: italic; }
QLabel#countBadge  { color: ${text_dim}; font-size: 10px; }
QFrame[frameShape="4"] { color: ${border_dim}; }
QFrame[frameShape="5"] { color: ${border_dim}; }
"""

def _build_style(palette: dict) -> str:
    return _re.sub(r"\$\{(\w+)\}", lambda m: palette.get(m.group(1), ""), _CSS)

DARK_STYLE  = _build_style(_DARK)
LIGHT_STYLE = _build_style(_LIGHT)


# ===========================================================================
# Background threads
# ===========================================================================

# ===========================================================================
# Output parsing  (Qt-free, unit-testable in isolation)
# ===========================================================================

def flatten_records(records, cols, rows, depth=0):
    """Flatten Volatility's nested ``__children`` tree records into flat rows,
    indenting the first column by depth to preserve hierarchy visually."""
    for rec in records:
        row = []
        for i, col in enumerate(cols):
            val = str(rec.get(col, "") or "")
            if i == 0 and depth:
                val = "  " * depth + val
            row.append(val)
        rows.append(row)
        for child in rec.get("__children", []):
            flatten_records([child], cols, rows, depth + 1)


def json_to_table(data):
    """Convert decoded Volatility JSON output into ``(columns, rows)``."""
    if isinstance(data, list) and data and isinstance(data[0], dict):
        cols = [k for k in data[0].keys() if k != "__children"]
        rows: List[List[str]] = []
        flatten_records(data, cols, rows)
        return cols, rows
    if isinstance(data, list) and data and isinstance(data[0], list):
        return [str(c) for c in data[0]], [[str(v) for v in r] for r in data[1:]]
    if isinstance(data, dict):
        if "columns" in data and "rows" in data:
            return ([str(c) for c in data["columns"]],
                    [[str(v) for v in r] for r in data["rows"]])
        cols = list(data.keys())
        return cols, [[str(data[c]) for c in cols]]
    return ["Result"], [[str(data)]]


def parse_vol_output(raw):
    """Parse raw stdout from ``vol.py --renderer json`` into ``(columns, rows)``.

    Handles three shapes: a single JSON document, newline-delimited JSON
    objects (JSONL), and—failing both—plain text echoed one line per row.
    """
    raw = (raw or "").strip()
    if not raw:
        return [], []
    try:
        return json_to_table(json.loads(raw))
    except json.JSONDecodeError:
        pass
    records = []
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
            if isinstance(obj, dict):
                records.append(obj)
        except json.JSONDecodeError:
            pass
    if records:
        return json_to_table(records)
    return ["Output"], [[l] for l in raw.splitlines() if l.strip()]


# ===========================================================================
# Worker threads
# ===========================================================================

class PluginDiscoveryThread(QThread):
    plugins_ready = pyqtSignal(dict)
    error         = pyqtSignal(str)

    def run(self):
        try:
            vol_framework.import_files(vol_plugins_module, ignore_errors=True)
            raw  = vol_framework.list_plugins()
            cats = {"windows": {}, "linux": {}, "mac": {}, "other": {}}
            for name, cls in raw.items():
                top = name.split(".")[0].lower()
                cats[top if top in cats else "other"][name] = cls
            self.plugins_ready.emit(cats)
        except Exception as exc:
            self.error.emit(str(exc))


class PluginRunnerThread(QThread):
    log_line    = pyqtSignal(str, str)
    result_data = pyqtSignal(str, str, list, list)
    finished    = pyqtSignal(bool)

    def __init__(self, plugin_name, dump_path, plugin_args, output_dir="",
                 symbol_paths=None, parent=None):
        super().__init__(parent)
        self.plugin_name  = plugin_name
        self.dump_path    = dump_path
        self.plugin_args  = plugin_args
        self.output_dir   = output_dir
        self.symbol_paths = symbol_paths or {}   # {"linux": [...], "mac": [...]}
        self._proc        = None

    def abort(self):
        if self._proc and self._proc.poll() is None:
            self._proc.terminate()

    def run(self):
        cmd = [sys.executable, str(SCRIPT_DIR / "vol.py"),
               "-q", "--renderer", "json", "-f", self.dump_path]
        # Inject custom symbol table paths so Linux/macOS plugins can resolve them
        for paths in self.symbol_paths.values():
            for p in paths:
                cmd.extend(["--symbols", p])
        if self.output_dir:
            cmd.extend(["-o", self.output_dir])
        cmd.append(self.plugin_name)

        for name, value in self.plugin_args.items():
            if value is None:
                continue
            flag = f"--{name}"
            if isinstance(value, bool):
                if value: cmd.append(flag)
            elif isinstance(value, list):
                if value: cmd.append(flag); cmd.extend(str(v) for v in value)
            elif str(value).strip():
                cmd.extend([flag, str(value)])

        self.log_line.emit(f"$ {' '.join(cmd)}", "cmd")
        try:
            self._proc = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                text=True, bufsize=1, cwd=str(SCRIPT_DIR))
            # Drain stderr live on a helper thread so Volatility's progress
            # lines surface in the log as they happen, not in one burst at the
            # end. stdout is read to completion on this thread and parsed once
            # the (single JSON document) output is fully available.
            stderr_thread = threading.Thread(
                target=self._drain_stderr, args=(self._proc.stderr,), daemon=True)
            stderr_thread.start()
            stdout = self._proc.stdout.read()
            self._proc.wait()
            stderr_thread.join(timeout=2)
            cols, rows = parse_vol_output(stdout)
            self.result_data.emit(self.plugin_name, self.dump_path, cols, rows)
            self.finished.emit(self._proc.returncode == 0)
        except Exception as exc:
            self.log_line.emit(f"Runner error: {exc}", "error")
            self.finished.emit(False)

    def _drain_stderr(self, stream):
        """Emit each stderr line as it arrives (runs on a helper thread)."""
        try:
            for line in iter(stream.readline, ""):
                line = line.rstrip("\n")
                if line.strip():
                    self.log_line.emit(
                        line, "error" if "error" in line.lower() else "debug")
        except (OSError, ValueError):
            pass


class ImageHashThread(QThread):
    """Compute MD5 + SHA-256 of a memory image in one streamed pass.

    Hashing is essential for a defensible chain of custody, but full-image
    hashing of multi-GB dumps must never block the UI — hence a dedicated
    thread that reports coarse progress and the final digests.
    """
    progress = pyqtSignal(int)            # 0-100
    done     = pyqtSignal(dict)           # {"md5","sha256","size","algo_error"}

    _CHUNK = 4 * 1024 * 1024              # 4 MiB

    def __init__(self, path: str, parent=None):
        super().__init__(parent)
        self.path = path
        self._abort = False

    def abort(self):
        self._abort = True

    def run(self):
        try:
            size = os.path.getsize(self.path)
            md5  = hashlib.md5()
            sha  = hashlib.sha256()
            read = 0
            last_pct = -1
            with open(self.path, "rb") as fh:
                while True:
                    if self._abort:
                        return
                    chunk = fh.read(self._CHUNK)
                    if not chunk:
                        break
                    md5.update(chunk)
                    sha.update(chunk)
                    read += len(chunk)
                    pct = int(read * 100 / size) if size else 100
                    if pct != last_pct:
                        last_pct = pct
                        self.progress.emit(pct)
            self.done.emit({
                "md5":    md5.hexdigest(),
                "sha256": sha.hexdigest(),
                "size":   size,
            })
        except OSError as exc:
            self.done.emit({"md5": None, "sha256": None,
                            "size": 0, "algo_error": str(exc)})


# ===========================================================================
# Drop-zone widget
# ===========================================================================

class DropZoneWidget(QFrame):
    file_loaded = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("dropZone")
        self.setAcceptDrops(True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedHeight(52)
        self._build_ui()

    def _build_ui(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 0, 14, 0)
        lay.setSpacing(10)

        tag = QLabel("IMG")
        tag.setFixedWidth(36)
        tag.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tag.setStyleSheet(
            f"color:{_c()['accent']};font-size:9px;font-weight:700;"
            f"letter-spacing:1px;background:transparent;")

        self.main_lbl = QLabel("Drop a memory image here  —  or click to browse")
        self.main_lbl.setObjectName("dropMain")

        self.file_lbl = QLabel("")
        self.file_lbl.setObjectName("dropFile")
        self.file_lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        lay.addWidget(tag)
        lay.addWidget(self.main_lbl)
        lay.addStretch()
        lay.addWidget(self.file_lbl)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            c = _c()
            self.setStyleSheet(
                f"QFrame#dropZone{{border:1px solid {c['accent']};"
                f"border-left:3px solid {c['accent']};background:{c['bg_hover']};}}"
            )

    def dragLeaveEvent(self, event):
        self.setStyleSheet("")

    def dropEvent(self, event):
        self.setStyleSheet("")
        urls = event.mimeData().urls()
        if urls:
            self._load(urls[0].toLocalFile())

    def mousePressEvent(self, event):
        settings = QSettings("vol3gui", "prefs")
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Memory Image", settings.value("last_dir", ""),
            "Memory Images (*.dmp *.mem *.vmem *.raw *.img *.bin *.lime *.dd *.E01 *.e01);;"
            "All Files (*)")
        if path:
            self._load(path)

    def _load(self, path: str):
        if not os.path.isfile(path):
            return
        QSettings("vol3gui", "prefs").setValue("last_dir", str(Path(path).parent))
        size = os.path.getsize(path)
        self.file_lbl.setText(f"✓  {os.path.basename(path)}  ({_fmt_size(size)})")
        self.main_lbl.setText("Image loaded  —  drop another to replace")
        self.file_loaded.emit(path)


# ===========================================================================
# Plugin browser
# ===========================================================================

class PluginBrowserWidget(QWidget):
    plugin_selected   = pyqtSignal(str, object)
    run_requested     = pyqtSignal(str, object)
    batch_run_requested = pyqtSignal(list)      # [(name, cls), …]
    profile_applied   = pyqtSignal(str, int)    # (profile name, #checked)

    _CATS = {"windows": "WINDOWS", "linux": "LINUX", "mac": "MACOS", "other": "OTHER"}

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("sidePanel")
        self._all: List[Tuple[str, Any, QTreeWidgetItem]] = []
        self._mitre_techs: Optional[frozenset] = None   # None = no filter
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── section header ─────────────────────────────────────────────────
        hdr = QWidget()
        hdr.setObjectName("subHeader")
        hdr.setFixedHeight(34)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(10, 0, 10, 0)
        lbl = QLabel("PLUGINS")
        lbl.setObjectName("sectionLabel")
        self.count_lbl = QLabel("")
        self.count_lbl.setObjectName("countBadge")
        hl.addWidget(lbl)
        hl.addStretch()
        hl.addWidget(self.count_lbl)
        lay.addWidget(hdr)

        # ── search ─────────────────────────────────────────────────────────
        sw = QWidget()
        sw.setObjectName("subHeader")
        sl = QHBoxLayout(sw)
        sl.setContentsMargins(8, 6, 8, 6)
        self.search = QLineEdit()
        self.search.setPlaceholderText("Filter plugins…")
        self.search.textChanged.connect(self._apply_filters)
        sl.addWidget(self.search)
        lay.addWidget(sw)

        # ── MITRE / threat-actor filter ────────────────────────────────────
        mw = QWidget()
        mw.setObjectName("subHeader")
        ml = QVBoxLayout(mw)
        ml.setContentsMargins(8, 4, 8, 6)
        ml.setSpacing(3)
        mitre_hdr = QLabel("MITRE / THREAT ACTOR FILTER")
        mitre_hdr.setObjectName("sectionLabel")
        ml.addWidget(mitre_hdr)
        self.mitre_combo = QComboBox()
        self.mitre_combo.setToolTip(
            "Filter the plugin list by MITRE ATT&CK Technique ID or known threat actor.\n"
            "Only plugins that map to the selected technique(s) will be shown.")
        self._populate_mitre_combo()
        self.mitre_combo.currentIndexChanged.connect(self._on_mitre_changed)
        ml.addWidget(self.mitre_combo)
        lay.addWidget(mw)

        # ── batch / triage controls ────────────────────────────────────────
        bw = QWidget()
        bw.setObjectName("subHeader")
        bl = QVBoxLayout(bw)
        bl.setContentsMargins(8, 4, 8, 6)
        bl.setSpacing(3)
        batch_hdr = QLabel("TRIAGE / BATCH")
        batch_hdr.setObjectName("sectionLabel")
        bl.addWidget(batch_hdr)
        self.profile_combo = QComboBox()
        self.profile_combo.setToolTip(
            "Tick a triage profile to check its plugins, then Run Checked.")
        self.profile_combo.addItem("Apply triage profile…")
        for pname in TRIAGE_PROFILES:
            self.profile_combo.addItem(pname)
        self.profile_combo.activated.connect(self._on_profile_chosen)
        bl.addWidget(self.profile_combo)
        brow = QHBoxLayout()
        brow.setSpacing(4)
        self.run_checked_btn = _ghost_btn("Run Checked (0)")
        self.run_checked_btn.setEnabled(False)
        self.run_checked_btn.clicked.connect(self._emit_batch)
        clr_btn = _ghost_btn("Clear")
        clr_btn.setFixedWidth(54)
        clr_btn.clicked.connect(self._clear_checks)
        brow.addWidget(self.run_checked_btn)
        brow.addWidget(clr_btn)
        bl.addLayout(brow)
        lay.addWidget(bw)

        # ── tree ────────────────────────────────────────────────────────────
        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(True)
        self.tree.setRootIsDecorated(True)
        self.tree.setAnimated(False)
        self.tree.setIndentation(16)
        self.tree.itemSelectionChanged.connect(self._on_select)
        self.tree.itemDoubleClicked.connect(self._on_dbl)
        self.tree.itemChanged.connect(self._on_item_changed)
        self.tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._ctx)
        lay.addWidget(self.tree)

    def _populate_mitre_combo(self):
        """Fill the MITRE / threat-actor combo box."""
        cb = self.mitre_combo
        cb.blockSignals(True)
        cb.clear()

        cb.addItem("— All Plugins —")
        cb.setItemData(0, None)   # sentinel: no filter

        # ── Technique IDs ──────────────────────────────────────────────────
        hdr_idx = cb.count()
        cb.addItem("── MITRE ATT&CK Techniques ──")
        cb.model().item(hdr_idx).setEnabled(False)

        for tid in sorted(MITRE_TECHNIQUES.keys()):
            label = f"{tid}  –  {MITRE_TECHNIQUES[tid]}"
            cb.addItem(label)
            cb.setItemData(cb.count() - 1, frozenset({tid}))

        # ── Threat actors ──────────────────────────────────────────────────
        hdr_idx2 = cb.count()
        cb.addItem("── Known Threat Actors / Groups ──")
        cb.model().item(hdr_idx2).setEnabled(False)

        for actor in sorted(THREAT_ACTORS.keys()):
            cb.addItem(actor)
            cb.setItemData(cb.count() - 1, frozenset(THREAT_ACTORS[actor]))

        cb.blockSignals(False)

    def _on_mitre_changed(self, _idx: int):
        data = self.mitre_combo.currentData()
        # Disabled header items return None; treat same as "All Plugins"
        if data is None and self.mitre_combo.currentIndex() != 0:
            # User clicked a disabled separator — jump back to "all"
            self.mitre_combo.blockSignals(True)
            self.mitre_combo.setCurrentIndex(0)
            self.mitre_combo.blockSignals(False)
            data = None
        self._mitre_techs = data
        self._apply_filters()

    def populate(self, cats: dict):
        self.tree.blockSignals(True)
        self.tree.clear()
        self._all.clear()
        total = 0
        for key in ("windows", "linux", "mac", "other"):
            plugins = cats.get(key, {})
            if not plugins: continue
            cat_item = QTreeWidgetItem([self._CATS.get(key, key)])
            f = cat_item.font(0)
            f.setWeight(QFont.Weight.Bold)
            f.setPointSize(9)
            cat_item.setFont(0, f)
            cat_item.setForeground(0, QColor(_c()["text_dim"]))
            cat_item.setFlags(cat_item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
            self.tree.addTopLevelItem(cat_item)
            for name, cls in sorted(plugins.items()):
                item = QTreeWidgetItem([name.split(".")[-1]])
                item.setData(0, Qt.ItemDataRole.UserRole, (name, cls))
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(0, Qt.CheckState.Unchecked)
                doc = (cls.__doc__ or "").strip()
                # Build tooltip: first line of docstring + MITRE techniques
                techs = _get_plugin_techniques(name)
                tip_parts = []
                if doc:
                    tip_parts.append(doc.split("\n")[0][:140])
                if techs:
                    tip_parts.append(
                        "MITRE: " + ", ".join(
                            f"{t} ({MITRE_TECHNIQUES.get(t, '')})" for t in techs
                        )
                    )
                if tip_parts:
                    item.setToolTip(0, "\n".join(tip_parts))
                cat_item.addChild(item)
                self._all.append((name, cls, item))
                total += 1
        self.count_lbl.setText(str(total))
        self.tree.blockSignals(False)
        self._update_batch_btn()
        self._apply_filters()

    def _apply_filters(self):
        """Apply both the text search and the MITRE/actor filter together."""
        text  = self.search.text().lower().strip()
        techs = self._mitre_techs   # frozenset of IDs, or None for "all"

        for name, _cls, item in self._all:
            text_ok = (not text) or (text in name.lower())

            if techs is None:
                mitre_ok = True
            else:
                plugin_techs = _get_plugin_techniques(name)
                # A plugin matches if any of its techniques starts-with (or
                # equals) any filter technique, supporting parent-level matching
                # (e.g. selecting T1003 also reveals T1003.001 plugins).
                mitre_ok = any(
                    pt.startswith(ft) or ft.startswith(pt)
                    for pt in plugin_techs
                    for ft in techs
                )

            item.setHidden(not (text_ok and mitre_ok))

        active = bool(text) or techs is not None
        vis_total = 0
        for i in range(self.tree.topLevelItemCount()):
            cat = self.tree.topLevelItem(i)
            vis = any(not cat.child(j).isHidden() for j in range(cat.childCount()))
            cat.setHidden(not vis)
            cat.setExpanded(active and vis)
            if vis:
                vis_total += sum(
                    1 for j in range(cat.childCount())
                    if not cat.child(j).isHidden()
                )

        total = len(self._all)
        self.count_lbl.setText(
            f"{vis_total}/{total}" if active else str(total)
        )

    def _on_select(self):
        items = self.tree.selectedItems()
        if items:
            data = items[0].data(0, Qt.ItemDataRole.UserRole)
            if data: self.plugin_selected.emit(*data)

    def _on_dbl(self, item: QTreeWidgetItem, _col: int):
        data = item.data(0, Qt.ItemDataRole.UserRole)
        if data: self.run_requested.emit(*data)

    def _ctx(self, pos):
        item = self.tree.itemAt(pos)
        if not item: return
        data = item.data(0, Qt.ItemDataRole.UserRole)
        if not data: return
        menu = QMenu(self)
        menu.addAction("Configure", lambda: self.plugin_selected.emit(*data))
        menu.addAction("Run with defaults", lambda: self.run_requested.emit(*data))
        menu.exec(self.tree.viewport().mapToGlobal(pos))

    # ── batch / triage ─────────────────────────────────────────────────────

    def _on_item_changed(self, _item, _col):
        self._update_batch_btn()

    def _checked_plugins(self) -> List[Tuple[str, Any]]:
        out = []
        for name, cls, item in self._all:
            if item.checkState(0) == Qt.CheckState.Checked:
                out.append((name, cls))
        return out

    def _update_batch_btn(self):
        n = len(self._checked_plugins())
        self.run_checked_btn.setText(f"Run Checked ({n})")
        self.run_checked_btn.setEnabled(n > 0)

    def _clear_checks(self):
        self.tree.blockSignals(True)
        for _name, _cls, item in self._all:
            item.setCheckState(0, Qt.CheckState.Unchecked)
        self.tree.blockSignals(False)
        self._update_batch_btn()

    def _on_profile_chosen(self, idx: int):
        if idx <= 0:
            return
        profile = self.profile_combo.itemText(idx)
        entries = TRIAGE_PROFILES.get(profile, [])
        self.tree.blockSignals(True)
        matched = 0
        # Applying a profile replaces the current selection rather than adding
        # to it, so the checked count always reflects just this profile.
        for name, _cls, item in self._all:
            if plugin_in_profile(name, entries):
                item.setCheckState(0, Qt.CheckState.Checked)
                matched += 1
            else:
                item.setCheckState(0, Qt.CheckState.Unchecked)
        self.tree.blockSignals(False)
        # Reset the combo back to its prompt row (it behaves as an action menu)
        self.profile_combo.blockSignals(True)
        self.profile_combo.setCurrentIndex(0)
        self.profile_combo.blockSignals(False)
        self._update_batch_btn()
        self.profile_applied.emit(profile, matched)

    def _emit_batch(self):
        plugins = self._checked_plugins()
        if plugins:
            self.batch_run_requested.emit(plugins)

    def restyle(self):
        """Re-apply theme-dependent colours (called on theme toggle)."""
        c = _c()
        for i in range(self.tree.topLevelItemCount()):
            self.tree.topLevelItem(i).setForeground(0, QColor(c["text_dim"]))


# ===========================================================================
# Plugin configuration panel
# ===========================================================================

class PluginConfigPanel(QWidget):
    run_requested = pyqtSignal(str, object, dict, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._plugin_name: Optional[str] = None
        self._plugin_cls  = None
        self._fields: Dict[str, Tuple[Any, QWidget]] = {}
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # section header
        hdr = QWidget()
        hdr.setObjectName("subHeader")
        hdr.setFixedHeight(34)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(10, 0, 10, 0)
        lbl = QLabel("CONFIGURE")
        lbl.setObjectName("sectionLabel")
        hl.addWidget(lbl)
        lay.addWidget(hdr)

        # plugin info block
        info_w = QWidget()
        info_w.setObjectName("pluginInfo")
        info_w.setMinimumHeight(60)
        info_l = QVBoxLayout(info_w)
        info_l.setContentsMargins(12, 10, 12, 10)
        info_l.setSpacing(4)
        self.name_lbl = QLabel("Select a plugin")
        self.name_lbl.setObjectName("pluginTitle")
        self.desc_lbl = QLabel("Choose a plugin from the panel on the left")
        self.desc_lbl.setObjectName("pluginDesc")
        self.desc_lbl.setWordWrap(True)
        info_l.addWidget(self.name_lbl)
        info_l.addWidget(self.desc_lbl)
        lay.addWidget(info_w)

        # options scroll
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.form_w = QWidget()
        self.form   = QFormLayout(self.form_w)
        self.form.setSpacing(8)
        self.form.setContentsMargins(12, 10, 12, 10)
        self.form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)
        scroll.setWidget(self.form_w)
        lay.addWidget(scroll, 1)

        # output dir
        od_w = QWidget()
        od_w.setObjectName("configFooter")
        od_l = QVBoxLayout(od_w)
        od_l.setContentsMargins(12, 8, 12, 8)
        od_l.setSpacing(4)
        od_hdr = QLabel("OUTPUT DIRECTORY")
        od_hdr.setObjectName("sectionLabel")
        od_row = QHBoxLayout()
        self.out_edit = QLineEdit()
        self.out_edit.setPlaceholderText("Optional – for dumped files")
        out_btn = _ghost_btn("Browse")
        out_btn.setMaximumWidth(62)
        out_btn.clicked.connect(self._browse_outdir)
        od_row.addWidget(self.out_edit)
        od_row.addWidget(out_btn)
        od_l.addWidget(od_hdr)
        od_l.addLayout(od_row)
        lay.addWidget(od_w)

        # run button
        run_w = QWidget()
        run_w.setObjectName("configFooter")
        rl = QHBoxLayout(run_w)
        rl.setContentsMargins(12, 10, 12, 10)
        self.run_btn = QPushButton("▶   Run Plugin")
        self.run_btn.setEnabled(False)
        self.run_btn.setMinimumHeight(36)
        self.run_btn.clicked.connect(self._emit_run)
        rl.addWidget(self.run_btn)
        lay.addWidget(run_w)

    def load_plugin(self, name: str, cls):
        self._plugin_name = name
        self._plugin_cls  = cls
        self._fields.clear()
        while self.form.count():
            item = self.form.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        self.name_lbl.setText(name.split(".")[-1])
        doc = " ".join((cls.__doc__ or "").split())
        self.desc_lbl.setText((doc[:220] + "…") if len(doc) > 220 else (doc or "No description."))

        try:
            reqs = cls.get_requirements()
        except Exception:
            reqs = []

        shown = 0
        for req in reqs:
            rtype = type(req).__name__
            if rtype in HIDDEN_REQ_TYPES or req.name in AUTO_FILLED_ARGS:
                continue
            widget = self._make_widget(req)
            if widget is None: continue
            lbl_text = req.name.replace("_", " ").replace("-", " ").title()
            if not getattr(req, "optional", True): lbl_text = "* " + lbl_text
            lbl = QLabel(lbl_text)
            lbl.setToolTip(getattr(req, "description", "") or "")
            self.form.addRow(lbl, widget)
            self._fields[req.name] = (req, widget)
            shown += 1

        if shown == 0:
            nl = QLabel("No additional options")
            nl.setObjectName("noneLabel")
            self.form.addRow(nl)

        self.run_btn.setEnabled(True)

    def enable_run(self, enabled: bool):
        self.run_btn.setEnabled(enabled and self._plugin_name is not None)

    def _make_widget(self, req) -> Optional[QWidget]:
        rtype   = type(req).__name__
        default = getattr(req, "default", None)

        if rtype == "BooleanRequirement":
            w = QCheckBox()
            if default: w.setChecked(bool(default))
            return w

        if rtype == "IntRequirement":
            w = QSpinBox()
            w.setRange(-(2**30), 2**30)
            if default is not None:
                try: w.setValue(int(default))
                except (TypeError, ValueError): pass
            return w

        if rtype == "ChoiceRequirement":
            w = QComboBox()
            if getattr(req, "optional", True): w.addItem("(any)")
            for c in getattr(req, "choices", []): w.addItem(str(c))
            if default is not None:
                idx = w.findText(str(default))
                if idx >= 0: w.setCurrentIndex(idx)
            return w

        if rtype == "URIRequirement":
            container = QWidget()
            h = QHBoxLayout(container)
            h.setContentsMargins(0, 0, 0, 0)
            h.setSpacing(4)
            edit = QLineEdit()
            edit.setPlaceholderText("file:///path/to/file")
            btn = _ghost_btn("…")
            btn.setFixedWidth(28)
            btn.clicked.connect(lambda: self._browse_uri(edit))
            h.addWidget(edit)
            h.addWidget(btn)
            container._edit = edit  # type: ignore[attr-defined]
            return container

        if rtype == "ListRequirement":
            elem = getattr(req, "element_type", str)
            w = QLineEdit()
            w.setPlaceholderText(f"Space-separated {getattr(elem,'__name__','values')}  (blank = all)")
            return w

        w = QLineEdit()
        if default is not None: w.setText(str(default))
        w.setPlaceholderText(getattr(req, "description", "") or req.name)
        return w

    def _browse_uri(self, edit: QLineEdit):
        path, _ = QFileDialog.getOpenFileName(self, "Select File", "")
        if path: edit.setText(f"file://{path}")

    def _browse_outdir(self):
        d = QFileDialog.getExistingDirectory(self, "Select Output Directory", "")
        if d: self.out_edit.setText(d)

    def _emit_run(self):
        if not self._plugin_name: return
        args: Dict[str, Any] = {}
        for req_name, (req, widget) in self._fields.items():
            val = self._read(req, widget)
            if val is not None: args[req_name] = val
        self.run_requested.emit(self._plugin_name, self._plugin_cls, args, self.out_edit.text().strip())

    def _read(self, req, widget) -> Any:
        rtype = type(req).__name__
        if rtype == "BooleanRequirement": return widget.isChecked()
        if rtype == "IntRequirement":
            val = widget.value()
            return None if (getattr(req, "optional", True) and val == widget.minimum()) else val
        if rtype == "ChoiceRequirement":
            txt = widget.currentText()
            return None if txt == "(any)" else txt
        if rtype == "URIRequirement":
            edit = getattr(widget, "_edit", None)
            txt = (edit.text() if edit else "").strip()
            return txt or None
        if rtype == "ListRequirement":
            txt = widget.text().strip()
            return txt.split() if txt else None
        txt = getattr(widget, "text", lambda: "")().strip()
        return txt or None


# ===========================================================================
# Results tab
# ===========================================================================

class ResultsTab(QWidget):
    def __init__(self, plugin_name, dump_path, columns, rows, parent=None,
                 image_meta=None, ts=None):
        super().__init__(parent)
        self.plugin_name = plugin_name
        self.dump_path   = dump_path
        self.columns     = columns
        self.rows        = rows
        # Snapshot of the image's chain-of-custody hashes at result time
        self.image_meta  = dict(image_meta) if image_meta else {}
        self._ts         = ts or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._build_ui()

    def _image_hash_lines(self) -> List[str]:
        """Audit lines (label : value) describing the source image's hashes."""
        lines = []
        sha = self.image_meta.get("sha256")
        md5 = self.image_meta.get("md5")
        if sha:
            lines.append(("Image SHA-256", sha))
        if md5:
            lines.append(("Image MD5", md5))
        return lines

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── toolbar ────────────────────────────────────────────────────────
        tb_w = QWidget()
        tb_w.setObjectName("tabToolbar")
        tb_w.setFixedHeight(40)
        tb = QHBoxLayout(tb_w)
        tb.setContentsMargins(10, 0, 10, 0)
        tb.setSpacing(6)

        self.info_lbl = QLabel()
        self.info_lbl.setObjectName("tabInfo")
        self._refresh_info_lbl()
        tb.addWidget(self.info_lbl)
        tb.addStretch()

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Filter rows…")
        self.search_box.setFixedWidth(180)
        self.search_box.textChanged.connect(self._filter)
        tb.addWidget(self.search_box)

        col_btn = _ghost_btn("Columns ▾")
        col_btn.setFixedWidth(90)
        col_btn.clicked.connect(self._column_picker)
        tb.addWidget(col_btn)

        exp_btn = _ghost_btn("Export ▾")
        exp_btn.setFixedWidth(80)
        exp_btn.clicked.connect(self._export_menu)
        tb.addWidget(exp_btn)

        lay.addWidget(tb_w)

        # ── table ──────────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setShowGrid(False)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._ctx_menu)
        self._populate()
        lay.addWidget(self.table)

        # ── footer ─────────────────────────────────────────────────────────
        ft_w = QWidget()
        ft_w.setObjectName("tabFooter")
        ft_w.setFixedHeight(24)
        ft = QHBoxLayout(ft_w)
        ft.setContentsMargins(10, 0, 10, 0)
        self._count_lbl = QLabel(f"{len(self.rows):,} rows")
        self._count_lbl.setObjectName("tabCount")
        ts_lbl = QLabel(self._ts)
        ts_lbl.setObjectName("tabTs")
        ft.addWidget(self._count_lbl)
        ft.addStretch()
        ft.addWidget(ts_lbl)
        lay.addWidget(ft_w)

    def _refresh_info_lbl(self):
        short = self.plugin_name.split(".")[-1]
        self.info_lbl.setText(
            f"{short}  ·  {len(self.rows):,} rows"
            f"  ·  {os.path.basename(self.dump_path)}"
        )

    def _populate(self):
        cols = self.columns or ["(no data)"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.setRowCount(len(self.rows))
        c = _c()
        clr_true = QColor(c["success"])
        clr_false = QColor(c["error"])
        clr_na   = QColor(c["text_muted"])
        for r, row in enumerate(self.rows):
            self.table.setRowHeight(r, 26)
            for col in range(len(cols)):
                val = row[col] if col < len(row) else ""
                item = QTableWidgetItem(val)
                vl = val.lower()
                if vl in ("true", "yes"):   item.setForeground(clr_true)
                elif vl in ("false", "no"): item.setForeground(clr_false)
                elif vl in ("n/a", "none", "-"): item.setForeground(clr_na)
                self.table.setItem(r, col, item)
        self.table.resizeColumnsToContents()
        for col in range(self.table.columnCount()):
            self.table.setColumnWidth(col, min(self.table.columnWidth(col), 360))

    def restyle(self):
        """Re-apply semantic cell colours when the theme changes."""
        if not self.rows: return
        c = _c()
        clr_true  = QColor(c["success"])
        clr_false = QColor(c["error"])
        clr_na    = QColor(c["text_muted"])
        clr_body  = QColor(c["text_body"])
        for r in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(r, col)
                if not item: continue
                vl = item.text().lower()
                if vl in ("true", "yes"):        item.setForeground(clr_true)
                elif vl in ("false", "no"):      item.setForeground(clr_false)
                elif vl in ("n/a", "none", "-"): item.setForeground(clr_na)
                else:                             item.setForeground(clr_body)

    def _filter(self, text: str):
        text = text.lower()
        visible = 0
        for r in range(self.table.rowCount()):
            row_text = " ".join(
                (self.table.item(r, c) or QTableWidgetItem("")).text().lower()
                for c in range(self.table.columnCount()))
            hidden = bool(text) and text not in row_text
            self.table.setRowHidden(r, hidden)
            if not hidden: visible += 1
        self._count_lbl.setText(
            f"{visible:,} / {len(self.rows):,} rows" if text else f"{len(self.rows):,} rows")

    def _column_picker(self):
        menu = QMenu(self)
        hdr = self.table.horizontalHeader()
        for col in range(self.table.columnCount()):
            lbl = self.table.horizontalHeaderItem(col).text()
            act = menu.addAction(lbl)
            act.setCheckable(True)
            act.setChecked(not hdr.isSectionHidden(col))
            act.toggled.connect(lambda checked, c=col: hdr.setSectionHidden(c, not checked))
        menu.exec(self.sender().mapToGlobal(self.sender().rect().bottomLeft()))

    def _ctx_menu(self, pos):
        item = self.table.itemAt(pos)
        if not item: return
        menu = QMenu(self)
        menu.addAction("Copy Cell",    lambda: QApplication.clipboard().setText(item.text()))
        menu.addAction("Copy Row (TSV)", lambda: QApplication.clipboard().setText(self._row_tsv(item.row())))
        menu.addSeparator()
        self._add_export_actions(menu)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _row_tsv(self, r: int) -> str:
        return "\t".join(
            (self.table.item(r, c) or QTableWidgetItem("")).text()
            for c in range(self.table.columnCount()))

    # ── exports ───────────────────────────────────────────────────────────

    def _export_menu(self):
        menu = QMenu(self)
        self._add_export_actions(menu)
        menu.exec(self.sender().mapToGlobal(self.sender().rect().bottomLeft()))

    def _add_export_actions(self, menu: QMenu):
        menu.addAction("CSV (.csv)",           self._export_csv)
        menu.addAction("TSV (.tsv)",           self._export_tsv)
        menu.addAction("JSON (.json)",         self._export_json)
        menu.addAction("Plain text (.txt)",    self._export_txt)
        menu.addAction("HTML report (.html)",  self._export_html)
        if HAS_PRINT:
            menu.addAction("PDF (.pdf)",       self._export_pdf)
        else:
            menu.addAction("PDF  (Qt print unavailable)").setEnabled(False)
        if HAS_XLSX:
            menu.addAction("Excel (.xlsx)",    self._export_xlsx)
        else:
            menu.addAction("Excel  (pip install openpyxl)").setEnabled(False)

    def _visible_data(self) -> Tuple[List[str], List[List[str]]]:
        hdr = self.table.horizontalHeader()
        col_idx = [c for c in range(self.table.columnCount()) if not hdr.isSectionHidden(c)]
        cols = [self.table.horizontalHeaderItem(c).text() for c in col_idx]
        rows = []
        for r in range(self.table.rowCount()):
            if not self.table.isRowHidden(r):
                rows.append([(self.table.item(r, c) or QTableWidgetItem("")).text() for c in col_idx])
        return cols, rows

    def _stem(self) -> str:
        return f"{self.plugin_name.replace('.','_')}_{_ts()}"

    # ── MITRE metadata helpers ─────────────────────────────────────────────

    def _mitre_meta(self) -> dict:
        """Return MITRE technique and threat-actor metadata for this plugin."""
        techs = _get_plugin_techniques(self.plugin_name)
        tech_labels = [
            f"{t} – {MITRE_TECHNIQUES.get(t, t)}" for t in techs
        ]
        # Find actors whose technique sets overlap this plugin's techniques
        matched_actors = []
        for actor, actor_techs in sorted(THREAT_ACTORS.items()):
            actor_set = set(actor_techs)
            if any(
                pt.startswith(ft) or ft.startswith(pt)
                for pt in techs for ft in actor_set
            ):
                matched_actors.append(actor)
        return {
            "technique_ids":    techs,
            "technique_labels": tech_labels,
            "threat_actors":    matched_actors,
        }

    # ── exports ───────────────────────────────────────────────────────────

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export CSV", f"{self._stem()}.csv", "CSV (*.csv)")
        if not path: return
        cols, rows = self._visible_data()
        meta = self._mitre_meta()
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            # MITRE header comments
            w.writerow([f"# Plugin: {self.plugin_name}"])
            w.writerow([f"# Image: {os.path.basename(self.dump_path)}"])
            for _hlabel, _hval in self._image_hash_lines():
                w.writerow([f"# {_hlabel}: {_hval}"])
            w.writerow([f"# Timestamp: {self._ts}"])
            if meta["technique_labels"]:
                w.writerow([f"# MITRE Techniques: {'; '.join(meta['technique_labels'])}"])
            if meta["threat_actors"]:
                w.writerow([f"# Threat Actors: {'; '.join(meta['threat_actors'])}"])
            w.writerow([])
            w.writerow(cols)
            w.writerows(rows)
        _done(self, path)

    def _export_tsv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export TSV", f"{self._stem()}.tsv", "TSV (*.tsv)")
        if not path: return
        cols, rows = self._visible_data()
        meta = self._mitre_meta()
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, delimiter="\t")
            w.writerow([f"# Plugin: {self.plugin_name}"])
            w.writerow([f"# Image: {os.path.basename(self.dump_path)}"])
            for _hlabel, _hval in self._image_hash_lines():
                w.writerow([f"# {_hlabel}: {_hval}"])
            w.writerow([f"# Timestamp: {self._ts}"])
            if meta["technique_labels"]:
                w.writerow([f"# MITRE Techniques: {'; '.join(meta['technique_labels'])}"])
            if meta["threat_actors"]:
                w.writerow([f"# Threat Actors: {'; '.join(meta['threat_actors'])}"])
            w.writerow([])
            w.writerow(cols)
            w.writerows(rows)
        _done(self, path)

    def _export_json(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export JSON", f"{self._stem()}.json", "JSON (*.json)")
        if not path: return
        cols, rows = self._visible_data()
        meta = self._mitre_meta()
        data = {
            "plugin":    self.plugin_name,
            "image":     self.dump_path,
            "image_meta": {
                "sha256": self.image_meta.get("sha256"),
                "md5":    self.image_meta.get("md5"),
                "size":   self.image_meta.get("size"),
            },
            "timestamp": self._ts,
            "mitre": {
                "techniques": [
                    {"id": t, "name": MITRE_TECHNIQUES.get(t, t)}
                    for t in meta["technique_ids"]
                ],
                "threat_actors": meta["threat_actors"],
            },
            "rows": [dict(zip(cols, r)) for r in rows],
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
        _done(self, path)

    def _export_txt(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export Text", f"{self._stem()}.txt", "Text (*.txt)")
        if not path: return
        cols, rows = self._visible_data()
        meta = self._mitre_meta()
        widths = [max(len(c), max((len(r[i]) for r in rows), default=0)) for i, c in enumerate(cols)]
        sep = "  ".join("-" * w for w in widths)
        hdr = "  ".join(c.ljust(widths[i]) for i, c in enumerate(cols))
        lines = [
            f"Plugin    : {self.plugin_name}",
            f"Image     : {self.dump_path}",
        ]
        for _hlabel, _hval in self._image_hash_lines():
            lines.append(f"{_hlabel.replace('Image ', '').ljust(10)}: {_hval}")
        lines += [
            f"Timestamp : {self._ts}",
            f"Rows      : {len(rows):,}",
        ]
        if meta["technique_labels"]:
            # Wrap long technique lists at 100 chars
            tline = "; ".join(meta["technique_labels"])
            lines.append(f"MITRE     : {tline}")
        if meta["threat_actors"]:
            # Split actors across lines if there are many
            actors = meta["threat_actors"]
            lines.append(f"Actors    : {'; '.join(actors[:4])}")
            for chunk in [actors[i:i+4] for i in range(4, len(actors), 4)]:
                lines.append(f"            {'; '.join(chunk)}")
        lines += ["", hdr, sep]
        for row in rows:
            lines.append("  ".join((row[i] if i < len(row) else "").ljust(widths[i]) for i in range(len(cols))))
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        _done(self, path)

    def _export_html(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export HTML", f"{self._stem()}.html", "HTML (*.html)")
        if not path: return
        cols, rows = self._visible_data()
        meta  = self._mitre_meta()
        th    = "".join(f"<th>{_esc(c)}</th>" for c in cols)
        tbody = "".join("<tr>" + "".join(f"<td>{_esc(v)}</td>" for v in r) + "</tr>\n" for r in rows)

        # MITRE badge block
        tech_badges = "".join(
            f'<span class="badge-tech" title="{_esc(MITRE_TECHNIQUES.get(t, t))}">{_esc(t)}</span>'
            for t in meta["technique_ids"]
        )
        actor_badges = "".join(
            f'<span class="badge-actor">{_esc(a)}</span>'
            for a in meta["threat_actors"]
        )
        hash_block = ""
        hlines = self._image_hash_lines()
        if hlines:
            hash_block = '<div class="meta" style="font-family:Consolas,monospace">' + \
                " &nbsp;·&nbsp; ".join(
                    f"{_esc(lbl)}: {_esc(val)}" for lbl, val in hlines) + "</div>"

        mitre_block = ""
        if tech_badges or actor_badges:
            mitre_block = (
                '<div class="mitre-section">'
                + ('<div class="mitre-row"><span class="mitre-lbl">MITRE ATT&amp;CK</span>'
                   + tech_badges + '</div>' if tech_badges else "")
                + ('<div class="mitre-row"><span class="mitre-lbl">Threat Actors</span>'
                   + actor_badges + '</div>' if actor_badges else "")
                + '</div>'
            )

        html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>Vol3 – {_esc(self.plugin_name.split('.')[-1])}</title>
<style>
body{{background:#2b2d3e;color:#ced4f0;font-family:"Segoe UI",sans-serif;font-size:13px;padding:24px 32px}}
h1{{color:#e05472;font-size:18px;margin:0 0 4px}}
.meta{{color:#6870a8;font-size:11px;margin-bottom:12px}}
.mitre-section{{background:#232538;border:1px solid #3c4060;border-radius:4px;padding:10px 14px;margin-bottom:18px}}
.mitre-row{{display:flex;flex-wrap:wrap;align-items:center;gap:6px;margin-bottom:4px}}
.mitre-row:last-child{{margin-bottom:0}}
.mitre-lbl{{color:#6870a8;font-size:10px;font-weight:700;letter-spacing:.8px;text-transform:uppercase;min-width:110px}}
.badge-tech{{background:#3a4a8a;color:#ced4f0;border-radius:3px;padding:2px 8px;font-size:11px;
             font-family:Consolas,monospace;cursor:default}}
.badge-actor{{background:#4a2838;color:#e05472;border-radius:3px;padding:2px 8px;font-size:11px;cursor:default}}
.wrap{{overflow-x:auto}}
table{{border-collapse:collapse;width:100%;font-family:Consolas,monospace;font-size:12px}}
thead{{background:#232538;border-bottom:1px solid #3c4060}}
th{{padding:8px 12px;text-align:left;color:#6870a8;font-size:10px;font-weight:700;letter-spacing:.8px;text-transform:uppercase}}
tbody tr{{border-bottom:1px solid #2f3245}}
tbody tr:nth-child(even){{background:#2f3245}}
tbody tr:hover{{background:#3a4a8a}}
td{{padding:6px 12px}}
footer{{margin-top:16px;color:#3c4060;font-size:10px;text-align:right}}
</style></head>
<body>
<h1>{_esc(self.plugin_name.split('.')[-1])}</h1>
<div class="meta">{_esc(self.plugin_name)} &nbsp;·&nbsp; {_esc(os.path.basename(self.dump_path))} &nbsp;·&nbsp; {_esc(self._ts)} &nbsp;·&nbsp; {len(rows):,} rows</div>
{hash_block}
{mitre_block}
<div class="wrap"><table><thead><tr>{th}</tr></thead><tbody>{tbody}</tbody></table></div>
<footer>Generated by Remora · {_esc(self._ts)}</footer>
</body></html>"""
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        _done(self, path)

    def _export_pdf(self):
        if not HAS_PRINT:
            QMessageBox.warning(self, "Unavailable", "Qt print support not available.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export PDF", f"{self._stem()}.pdf", "PDF (*.pdf)")
        if not path: return
        cols, rows = self._visible_data()
        meta = self._mitre_meta()
        th    = "".join(f"<th>{_esc(c)}</th>" for c in cols)
        tbody = "".join("<tr>" + "".join(f"<td>{_esc(v)}</td>" for v in r) + "</tr>" for r in rows)

        mitre_rows = ""
        for _hlabel, _hval in self._image_hash_lines():
            mitre_rows += f"<tr><td><b>{_esc(_hlabel)}</b></td><td>{_esc(_hval)}</td></tr>"
        if meta["technique_labels"]:
            techs_html = ", ".join(_esc(t) for t in meta["technique_labels"])
            mitre_rows += f"<tr><td><b>MITRE Techniques</b></td><td>{techs_html}</td></tr>"
        if meta["threat_actors"]:
            actors_html = ", ".join(_esc(a) for a in meta["threat_actors"])
            mitre_rows += f"<tr><td><b>Threat Actors</b></td><td>{actors_html}</td></tr>"
        mitre_table = (
            f"<table border='0' cellpadding='3' style='font-size:7pt;margin-bottom:8px'>"
            f"{mitre_rows}</table>" if mitre_rows else ""
        )

        html = (
            f"<html><body style='font-family:sans-serif;font-size:9pt'>"
            f"<h3 style='margin-bottom:2px'>{_esc(self.plugin_name)}</h3>"
            f"<p style='font-size:8pt;color:#555;margin:0 0 6px'>"
            f"{_esc(self.dump_path)} · {_esc(self._ts)} · {len(rows):,} rows</p>"
            f"{mitre_table}"
            f"<table border='0' cellspacing='0' cellpadding='4'"
            f" style='border-collapse:collapse;font-size:8pt'>"
            f"<thead style='background:#dde;font-weight:bold'><tr>{th}</tr></thead>"
            f"<tbody>{tbody}</tbody></table></body></html>"
        )
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(path)
        printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        try: printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        except Exception: pass
        doc = QTextDocument()
        doc.setHtml(html)
        doc.print_(printer)
        _done(self, path)

    def _export_xlsx(self):
        if not HAS_XLSX:
            QMessageBox.warning(self, "Unavailable", "pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Excel", f"{self._stem()}.xlsx", "Excel (*.xlsx)")
        if not path: return
        cols, rows = self._visible_data()
        meta = self._mitre_meta()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.plugin_name.split(".")[-1][:31]

        # ── metadata rows ──────────────────────────────────────────────────
        mfont  = XLFont(bold=True, color="9AA2CC", size=10)
        vfont  = XLFont(color="CED4F0", size=10)
        mbg    = PatternFill("solid", fgColor="1E2030")

        def _meta_row(label, value):
            ws.append([label, value])
            r = ws.max_row
            for ci in (1, 2):
                ws.cell(r, ci).fill  = mbg
                ws.cell(r, ci).font  = mfont if ci == 1 else vfont

        _meta_row("Plugin",    self.plugin_name)
        _meta_row("Image",     self.dump_path)
        for _hlabel, _hval in self._image_hash_lines():
            _meta_row(_hlabel, _hval)
        _meta_row("Timestamp", self._ts)
        _meta_row("Rows",      len(rows))
        if meta["technique_labels"]:
            _meta_row("MITRE Techniques", "; ".join(meta["technique_labels"]))
        if meta["threat_actors"]:
            _meta_row("Threat Actors", "; ".join(meta["threat_actors"]))
        ws.append([])   # blank separator

        hrow = ws.max_row + 1
        hf    = PatternFill("solid", fgColor="232538")
        hfont = XLFont(bold=True, color="9AA2CC", size=10)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=hrow, column=ci, value=col)
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="left")

        f_odd  = PatternFill("solid", fgColor="2B2D3E")
        f_even = PatternFill("solid", fgColor="2F3245")
        dfont  = XLFont(name="Consolas", size=10, color="CED4F0")
        for ri, row in enumerate(rows, hrow + 1):
            fill = f_odd if ri % 2 else f_even
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill; cell.font = dfont

        # ── MITRE summary sheet ────────────────────────────────────────────
        if meta["technique_ids"] or meta["threat_actors"]:
            ms = wb.create_sheet("MITRE Coverage")
            ms.append(["Plugin", self.plugin_name])
            ms.append(["Timestamp", self._ts])
            ms.append([])
            ms.append(["Technique ID", "Technique Name"])
            for tid in meta["technique_ids"]:
                ms.append([tid, MITRE_TECHNIQUES.get(tid, "")])
            ms.append([])
            ms.append(["Threat Actors"])
            for actor in meta["threat_actors"]:
                ms.append([actor])

        for col_cells in ws.columns:
            ml = max((len(str(cell.value or "")) for cell in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(ml + 2, 50)
        wb.save(path)
        _done(self, path)


# ===========================================================================
# Log panel
# ===========================================================================

class LogPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._log_file: Optional[str] = None
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        hdr = QWidget()
        hdr.setObjectName("subHeader")
        hdr.setFixedHeight(28)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(10, 0, 10, 0)
        lbl = QLabel("LOG")
        lbl.setObjectName("sectionLabel")
        hl.addWidget(lbl)
        hl.addStretch()
        self._file_lbl = QLabel("")
        self._file_lbl.setObjectName("runLabel")
        self._file_lbl.setToolTip("Log file path")
        hl.addWidget(self._file_lbl)
        clr = _ghost_btn("Clear")
        clr.setFixedWidth(50)
        clr.setFixedHeight(22)
        clr.clicked.connect(self._clear)
        hl.addWidget(clr)
        lay.addWidget(hdr)

        self.text = QTextEdit()
        self.text.setReadOnly(True)
        lay.addWidget(self.text)

    def set_log_file(self, image_path: str):
        """Derive log filename from image stem + today's date and start appending."""
        stem = Path(image_path).stem
        # sanitise: replace characters that are awkward in filenames
        safe_stem = _re.sub(r'[^\w\-.]', '_', stem)
        date_str  = datetime.now().strftime("%Y-%m-%d")
        log_dir   = Path(image_path).parent
        self._log_file = str(log_dir / f"{safe_stem}_{date_str}.log")
        # Write a session-start separator so appended runs are visually distinct
        is_new = not Path(self._log_file).exists()
        with open(self._log_file, "a", encoding="utf-8") as fh:
            if not is_new:
                fh.write("\n")
            fh.write(
                f"{'='*72}\n"
                f"  Session started : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"  Image           : {image_path}\n"
                f"{'='*72}\n"
            )
        short = Path(self._log_file).name
        self._file_lbl.setText(f"→ {short}  ")
        self._file_lbl.setToolTip(self._log_file)

    def _clear(self):
        self.text.clear()

    def log(self, msg: str, level: str = "info"):
        c = _c()
        colours = {
            "cmd":     c["text_sec"],
            "info":    c["text_body"],
            "debug":   c["text_muted"],
            "success": c["success"],
            "warning": c["warning"],
            "error":   c["error"],
        }
        clr = colours.get(level, c["text_body"])
        now = datetime.now()
        ts_display = now.strftime("%H:%M:%S")
        self.text.append(
            f'<span style="color:{c["text_muted"]}">[{ts_display}]</span> '
            f'<span style="color:{clr}">{_esc(str(msg))}</span>'
        )
        sb = self.text.verticalScrollBar()
        sb.setValue(sb.maximum())

        # ── persist to log file ──────────────────────────────────────────
        if self._log_file:
            ts_full = now.strftime("%Y-%m-%d %H:%M:%S")
            try:
                with open(self._log_file, "a", encoding="utf-8") as fh:
                    fh.write(f"[{ts_full}] [{level.upper():<7}] {msg}\n")
            except OSError:
                pass  # don't crash the GUI over a log write failure


# ===========================================================================
# Volshell widget
# ===========================================================================

class VolshellWidget(QWidget):
    """Embedded volshell terminal – spawns volshell.py as a subprocess and
    streams its stdout/stderr into a read-only log area.  The user types
    commands in the input line which are written to the process stdin."""

    def __init__(self, dump_path: str, parent=None):
        super().__init__(parent)
        self.dump_path = dump_path
        self._proc: Optional[QProcess] = None
        self._build_ui()
        self._start()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── toolbar ────────────────────────────────────────────────────────
        tb = QWidget()
        tb.setObjectName("subHeader")
        tb.setFixedHeight(32)
        tl = QHBoxLayout(tb)
        tl.setContentsMargins(10, 0, 10, 0)
        tl.setSpacing(8)
        lbl = QLabel("VOLSHELL")
        lbl.setObjectName("sectionLabel")
        tl.addWidget(lbl)
        file_lbl = QLabel(f"  {os.path.basename(self.dump_path)}")
        file_lbl.setObjectName("runLabel")
        tl.addWidget(file_lbl)
        tl.addStretch()
        self._restart_btn = _ghost_btn("Restart")
        self._restart_btn.setFixedHeight(22)
        self._restart_btn.clicked.connect(self._restart)
        tl.addWidget(self._restart_btn)
        self._kill_btn = _ghost_btn("Kill")
        self._kill_btn.setFixedHeight(22)
        self._kill_btn.clicked.connect(self._kill)
        tl.addWidget(self._kill_btn)
        lay.addWidget(tb)

        # ── output area ───────────────────────────────────────────────────
        self.output = QTextEdit()
        self.output.setReadOnly(True)
        self.output.setObjectName("volshellOutput")
        font = QFont("Cascadia Code, Fira Code, JetBrains Mono, Consolas, monospace", 11)
        font.setStyleHint(QFont.StyleHint.Monospace)
        self.output.setFont(font)
        lay.addWidget(self.output, 1)

        # ── input bar ─────────────────────────────────────────────────────
        in_w = QWidget()
        in_w.setObjectName("configFooter")
        il = QHBoxLayout(in_w)
        il.setContentsMargins(8, 6, 8, 6)
        il.setSpacing(6)
        prompt = QLabel(">>>")
        prompt.setObjectName("runLabel")
        self.input = QLineEdit()
        self.input.setPlaceholderText("Type a Python expression and press Enter…")
        self.input.returnPressed.connect(self._send)
        send_btn = _ghost_btn("Send")
        send_btn.setFixedHeight(26)
        send_btn.clicked.connect(self._send)
        il.addWidget(prompt)
        il.addWidget(self.input, 1)
        il.addWidget(send_btn)
        lay.addWidget(in_w)

        # history
        self._history: List[str] = []
        self._hist_idx = -1
        self.input.installEventFilter(self)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        if obj is self.input and event.type() == QEvent.Type.KeyPress:
            from PyQt6.QtGui import QKeyEvent
            key = event.key()
            if key == Qt.Key.Key_Up:
                if self._history and self._hist_idx < len(self._history) - 1:
                    self._hist_idx += 1
                    self.input.setText(self._history[-(self._hist_idx + 1)])
                return True
            if key == Qt.Key.Key_Down:
                if self._hist_idx > 0:
                    self._hist_idx -= 1
                    self.input.setText(self._history[-(self._hist_idx + 1)])
                elif self._hist_idx == 0:
                    self._hist_idx = -1
                    self.input.clear()
                return True
        return super().eventFilter(obj, event)

    def _start(self):
        volshell = str(SCRIPT_DIR / "volshell.py")
        self._proc = QProcess(self)
        self._proc.setProcessChannelMode(QProcess.ProcessChannelMode.MergedChannels)
        self._proc.readyReadStandardOutput.connect(self._on_output)
        self._proc.finished.connect(self._on_finished)
        self._proc.started.connect(self._on_started)
        self._proc.errorOccurred.connect(self._on_error)
        self._proc.start(sys.executable, [volshell, "-f", self.dump_path])

    def _on_started(self):
        c = _c()
        self.output.append(
            f'<span style="color:{c["success"]}">● volshell started  —  '
            f'{os.path.basename(self.dump_path)}</span>'
        )

    def _on_output(self):
        raw = bytes(self._proc.readAllStandardOutput()).decode("utf-8", errors="replace")
        c = _c()
        for line in raw.splitlines(keepends=True):
            self.output.moveCursor(QTextCursor.MoveOperation.End)
            self.output.insertPlainText(line)
        sb = self.output.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _on_finished(self, code, status):
        c = _c()
        self.output.append(
            f'<span style="color:{c["warning"]}">● volshell exited (code {code})</span>'
        )

    def _on_error(self, err):
        msgs = {
            QProcess.ProcessError.FailedToStart: "Failed to start volshell.py",
            QProcess.ProcessError.Crashed: "volshell process crashed",
            QProcess.ProcessError.Timedout: "Timeout",
            QProcess.ProcessError.ReadError: "Read error",
            QProcess.ProcessError.WriteError: "Write error",
        }
        c = _c()
        self.output.append(
            f'<span style="color:{c["error"]}">● {msgs.get(err, "Process error")}</span>'
        )

    def _send(self):
        cmd = self.input.text()
        if not cmd:
            return
        self._history.append(cmd)
        self._hist_idx = -1
        c = _c()
        self.output.append(
            f'<span style="color:{c["text_sec"]}">>>> {_esc(cmd)}</span>'
        )
        if self._proc and self._proc.state() == QProcess.ProcessState.Running:
            self._proc.write((cmd + "\n").encode())
        else:
            self.output.append(
                f'<span style="color:{c["error"]}">● volshell is not running — click Restart</span>'
            )
        self.input.clear()

    def _kill(self):
        if self._proc:
            self._proc.kill()

    def _restart(self):
        self._kill()
        if self._proc:
            self._proc.waitForFinished(2000)
            self._proc.deleteLater()
            self._proc = None
        self.output.clear()
        self._start()

    def closeEvent(self, event):
        self._kill()
        super().closeEvent(event)


# ===========================================================================
# MITRE Coverage Matrix dialog
# ===========================================================================

class MitreCoverageDialog(QWidget):
    """Standalone window: rows = plugins, columns = ATT&CK tactics.

    Each cell shows the highest confidence level (H / M / L) among the
    techniques the plugin maps to within that tactic column.
    H = plugin specifically designed for that detection
    M = strong secondary signal
    L = circumstantial / indirect
    Blank = no mapping.
    """

    _CONF_ORDER = {"H": 3, "M": 2, "L": 1, "": 0}
    _CONF_LABEL = {"H": "●  H", "M": "◉  M", "L": "○  L", "": ""}

    def __init__(self, plugin_cats: dict, parent=None):
        super().__init__(parent, Qt.WindowType.Window)
        self.setWindowTitle("MITRE ATT&CK Coverage Matrix  —  Remora")
        self.resize(1280, 820)
        self._plugin_cats = plugin_cats
        self._rows: List[Tuple[str, str, dict]] = []   # (full_name, short, tactic→conf)
        self._build_data()
        self._build_ui()

    # ── data ──────────────────────────────────────────────────────────────────

    def _build_data(self):
        """Pre-compute tactic → confidence for every plugin that has coverage."""
        tactics = list(MITRE_TACTICS.keys())
        for key in ("windows", "linux", "mac", "other"):
            for full_name in sorted(self._plugin_cats.get(key, {}).keys()):
                techs = _get_plugin_techniques(full_name)
                if not techs:
                    continue
                # For each tactic, find the best confidence among matching techniques
                tactic_conf: dict = {}
                plugin_parts = full_name.lower().split(".")
                # Identify which plugin_map key(s) fired so we can look up confidence
                fired_keys = []
                for pk in PLUGIN_MITRE_MAP:
                    pk_parts = pk.split(".")
                    klen = len(pk_parts)
                    for i in range(len(plugin_parts) - klen + 1):
                        if plugin_parts[i:i + klen] == pk_parts:
                            fired_keys.append(pk)
                            break

                for tactic in tactics:
                    tac_prefixes = MITRE_TACTICS[tactic]
                    best = ""
                    for tech in techs:
                        # Check if this technique belongs to this tactic
                        if not any(tech.startswith(p) or p.startswith(tech)
                                   for p in tac_prefixes):
                            continue
                        # Look up confidence from fired keys
                        for pk in fired_keys:
                            conf = _get_confidence(pk, tech)
                            if self._CONF_ORDER.get(conf, 0) > self._CONF_ORDER.get(best, 0):
                                best = conf
                    if best:
                        tactic_conf[tactic] = best

                if tactic_conf:
                    self._rows.append((full_name, full_name.split(".")[-1], tactic_conf))

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── toolbar ────────────────────────────────────────────────────────
        tb = QWidget()
        tb.setObjectName("subHeader")
        tb.setFixedHeight(42)
        tl = QHBoxLayout(tb)
        tl.setContentsMargins(12, 0, 12, 0)
        tl.setSpacing(8)

        lbl = QLabel("MITRE ATT&CK COVERAGE MATRIX")
        lbl.setObjectName("sectionLabel")
        tl.addWidget(lbl)

        legend = QLabel(
            "   ●  H = primary detection    ◉  M = secondary signal    ○  L = circumstantial")
        legend.setObjectName("tabInfo")
        tl.addWidget(legend)
        tl.addStretch()

        search = QLineEdit()
        search.setPlaceholderText("Filter plugins…")
        search.setFixedWidth(200)
        search.textChanged.connect(self._filter)
        self._search = search
        tl.addWidget(search)

        tactic_cb = QComboBox()
        tactic_cb.addItem("All Tactics")
        for t in MITRE_TACTICS:
            tactic_cb.addItem(t)
        tactic_cb.currentTextChanged.connect(self._on_tactic_filter)
        self._tactic_cb = tactic_cb
        tl.addWidget(tactic_cb)

        conf_cb = QComboBox()
        conf_cb.addItem("All Confidence")
        conf_cb.addItem("High only")
        conf_cb.addItem("High + Medium")
        conf_cb.currentTextChanged.connect(self._on_conf_filter)
        self._conf_cb = conf_cb
        tl.addWidget(conf_cb)

        exp_btn = _ghost_btn("Export CSV ▾")
        exp_btn.setFixedWidth(100)
        exp_btn.clicked.connect(self._export_csv)
        tl.addWidget(exp_btn)

        lay.addWidget(tb)

        # ── table ──────────────────────────────────────────────────────────
        tactics = list(MITRE_TACTICS.keys())
        self._tactics = tactics

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)

        headers = ["Plugin", "OS"] + tactics
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setStretchLastSection(False)

        self._populate_table(self._rows)
        lay.addWidget(self.table)

        # ── footer ─────────────────────────────────────────────────────────
        ft = QWidget()
        ft.setObjectName("tabFooter")
        ft.setFixedHeight(24)
        fl = QHBoxLayout(ft)
        fl.setContentsMargins(10, 0, 10, 0)
        self._count_lbl = QLabel()
        self._count_lbl.setObjectName("tabCount")
        self._count_lbl.setText(f"{len(self._rows)} plugins mapped")
        fl.addWidget(self._count_lbl)
        fl.addStretch()
        lay.addWidget(ft)

    def _populate_table(self, rows):
        c = _c()
        conf_colors = {
            "H": c["error"],     # red-pink — high severity / confidence
            "M": c["warning"],   # amber
            "L": c["text_sec"],  # muted
        }
        tactics = self._tactics

        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))

        os_map = {"windows": "Win", "linux": "Linux", "mac": "macOS",
                  "other": "Other"}

        for r, (full_name, short, tac_conf) in enumerate(rows):
            self.table.setRowHeight(r, 22)
            os_key = full_name.split(".")[0].lower()

            name_item = QTableWidgetItem(full_name)
            name_item.setToolTip(full_name)
            self.table.setItem(r, 0, name_item)

            os_item = QTableWidgetItem(os_map.get(os_key, os_key))
            os_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(r, 1, os_item)

            for col, tactic in enumerate(tactics, start=2):
                conf = tac_conf.get(tactic, "")
                label = self._CONF_LABEL.get(conf, "")
                cell = QTableWidgetItem(label)
                cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if conf in conf_colors:
                    cell.setForeground(QColor(conf_colors[conf]))
                # Store raw confidence for sorting (H=3, M=2, L=1, blank=0)
                cell.setData(Qt.ItemDataRole.UserRole, self._CONF_ORDER.get(conf, 0))
                self.table.setItem(r, col, cell)

        self.table.setSortingEnabled(True)
        self.table.resizeColumnToContents(0)
        self.table.setColumnWidth(0, min(self.table.columnWidth(0), 320))
        self.table.setColumnWidth(1, 56)
        for col in range(2, self.table.columnCount()):
            self.table.setColumnWidth(col, 90)

    def _filter(self, text: str):
        text = text.lower().strip()
        tactic = self._tactic_cb.currentText()
        conf_min = {"All Confidence": 0, "High only": 3, "High + Medium": 2}.get(
            self._conf_cb.currentText(), 0)
        tactic_cols = (
            [self._tactics.index(tactic) + 2]
            if tactic != "All Tactics" and tactic in self._tactics else
            list(range(2, 2 + len(self._tactics)))
        )
        visible = 0
        for r in range(self.table.rowCount()):
            name_item = self.table.item(r, 0)
            name = (name_item.text() if name_item else "").lower()
            text_ok = (not text) or text in name
            conf_ok = (conf_min == 0) or any(
                (self.table.item(r, c) or QTableWidgetItem("")).data(Qt.ItemDataRole.UserRole) >= conf_min
                for c in tactic_cols
            )
            hidden = not (text_ok and conf_ok)
            self.table.setRowHidden(r, hidden)
            if not hidden:
                visible += 1
        total = self.table.rowCount()
        self._count_lbl.setText(
            f"{visible}/{total} plugins" if (text or conf_min or tactic != "All Tactics")
            else f"{total} plugins mapped"
        )

    def _on_tactic_filter(self, _txt: str):
        self._filter(self._search.text())

    def _on_conf_filter(self, _txt: str):
        self._filter(self._search.text())

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Coverage Matrix", "mitre_coverage_matrix.csv", "CSV (*.csv)")
        if not path:
            return
        import csv as _csv
        tactics = self._tactics
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            w.writerow(["Plugin", "OS"] + tactics)
            for r in range(self.table.rowCount()):
                if self.table.isRowHidden(r):
                    continue
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(r, col)
                    # Export raw H/M/L instead of the symbol-label
                    if col >= 2:
                        rank = (item.data(Qt.ItemDataRole.UserRole) if item else 0) or 0
                        row_data.append({3: "H", 2: "M", 1: "L"}.get(rank, ""))
                    else:
                        row_data.append(item.text() if item else "")
                w.writerow(row_data)
        _done(self, path)

    def restyle(self):
        """Re-apply theme colours when palette changes."""
        c = _c()
        conf_colors = {"H": c["error"], "M": c["warning"], "L": c["text_sec"]}
        for r in range(self.table.rowCount()):
            for col in range(2, self.table.columnCount()):
                item = self.table.item(r, col)
                if not item:
                    continue
                rank = item.data(Qt.ItemDataRole.UserRole) or 0
                conf = {3: "H", 2: "M", 1: "L"}.get(rank, "")
                if conf:
                    item.setForeground(QColor(conf_colors[conf]))


# ===========================================================================
# Main window
# ===========================================================================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Remora  —  Memory Forensics")
        self.resize(1480, 920)

        self._dump_path: Optional[str] = None
        self._runner    = None
        self._tab_n     = 0
        # Evidence-integrity metadata for the loaded image (chain of custody)
        self._image_meta: dict = {}
        self._hash_thread: Optional[ImageHashThread] = None
        # Pending batch-run queue of (plugin_name, cls) tuples
        self._batch_queue: List[Tuple[str, Any]] = []
        self._batch_total: int = 0
        # custom symbol table paths: {"linux": [], "mac": []}
        self._symbol_paths: Dict[str, List[str]] = {"linux": [], "mac": []}
        self._volshell_tab_idx: int = -1
        self._coverage_win: Optional[MitreCoverageDialog] = None
        self._plugin_cats: dict = {}

        # Load saved theme before building UI
        global _ACTIVE
        settings = QSettings("vol3gui", "prefs")
        if not settings.value("dark_mode", True, type=bool):
            _ACTIVE = dict(_LIGHT)
            QApplication.instance().setStyleSheet(LIGHT_STYLE)
        else:
            _ACTIVE = dict(_DARK)
            QApplication.instance().setStyleSheet(DARK_STYLE)

        self._build_ui()
        self._build_menu()
        self._build_statusbar()
        if MITRE_OVERRIDE_STATUS:
            self.log_panel.log(MITRE_OVERRIDE_STATUS, "info")
        self._discover_plugins()

    # ── build ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        root_lay = QVBoxLayout(root)
        root_lay.setContentsMargins(0, 0, 0, 0)
        root_lay.setSpacing(0)

        root_lay.addWidget(self._make_header())

        self.drop_zone = DropZoneWidget()
        self.drop_zone.file_loaded.connect(self._on_dump_loaded)
        root_lay.addWidget(self.drop_zone)

        outer = QSplitter(Qt.Orientation.Horizontal)
        outer.setHandleWidth(1)

        self.browser = PluginBrowserWidget()
        self.browser.setMinimumWidth(190)
        self.browser.setMaximumWidth(300)
        self.browser.plugin_selected.connect(self._on_plugin_selected)
        self.browser.run_requested.connect(self._quick_run)
        self.browser.batch_run_requested.connect(self._on_batch_run)
        self.browser.profile_applied.connect(self._on_profile_applied)

        self.config_panel = PluginConfigPanel()
        self.config_panel.setMinimumWidth(260)
        self.config_panel.setMaximumWidth(400)
        self.config_panel.run_requested.connect(self._on_run)

        right = QSplitter(Qt.Orientation.Vertical)
        right.setHandleWidth(1)

        self.results_tabs = QTabWidget()
        self.results_tabs.setTabsClosable(True)
        self.results_tabs.setMovable(True)
        self.results_tabs.tabCloseRequested.connect(self._close_tab)
        self.results_tabs.tabCloseRequested.connect(self._on_tab_close_track)
        self._show_welcome()

        self.log_panel = LogPanel()
        self.log_panel.setMaximumHeight(180)
        self.log_panel.setMinimumHeight(72)

        right.addWidget(self.results_tabs)
        right.addWidget(self.log_panel)
        right.setSizes([680, 150])

        outer.addWidget(self.browser)
        outer.addWidget(self.config_panel)
        outer.addWidget(right)
        outer.setSizes([240, 330, 890])

        root_lay.addWidget(outer, 1)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setFixedHeight(2)
        self.progress.setTextVisible(False)
        root_lay.addWidget(self.progress)

    def _make_header(self) -> QWidget:
        w = QWidget()
        w.setObjectName("headerBar")
        w.setFixedHeight(40)
        lay = QHBoxLayout(w)
        lay.setContentsMargins(14, 0, 10, 0)
        lay.setSpacing(0)

        self._hdr_bar = QFrame()
        self._hdr_bar.setFixedWidth(3)
        self._hdr_bar.setFixedHeight(18)

        self.hdr_title = QLabel("  REMORA")
        self.hdr_sub   = QLabel("  /  memory forensics")

        lay.addWidget(self._hdr_bar)
        lay.addWidget(self.hdr_title)
        lay.addWidget(self.hdr_sub)
        lay.addStretch()

        self.hdr_file_chip = QLabel("no image loaded")
        self.hdr_file_chip.setObjectName("fileChip")
        lay.addWidget(self.hdr_file_chip)

        self.hdr_run_lbl = QLabel("")
        self.hdr_run_lbl.setObjectName("runLabel")
        lay.addWidget(self.hdr_run_lbl)

        self.theme_btn = QPushButton()
        self.theme_btn.setFixedSize(96, 24)
        self.theme_btn.clicked.connect(self._toggle_theme)
        lay.addSpacing(14)
        lay.addWidget(self.theme_btn)

        self._apply_header_colours()
        return w

    def _apply_header_colours(self):
        c = _c()
        is_dark = (_ACTIVE == _DARK or list(_ACTIVE.values()) == list(_DARK.values()))
        self._hdr_bar.setStyleSheet(f"background:{c['accent']};border:none;")
        self.hdr_title.setStyleSheet(
            f"color:{c['text_hi']};font-size:13px;font-weight:700;"
            f"letter-spacing:2px;background:transparent;")
        self.hdr_sub.setStyleSheet(
            f"color:{c['text_muted']};font-size:11px;background:transparent;")
        self.theme_btn.setText("LIGHT" if is_dark else "DARK")
        self.theme_btn.setStyleSheet(
            f"QPushButton{{background:transparent;color:{c['text_sec']};"
            f"border:1px solid {c['border']};border-radius:2px;"
            f"font-size:10px;font-weight:700;letter-spacing:1px;padding:0 8px;}}"
            f"QPushButton:hover{{background:{c['bg_hover']};color:{c['text_body']};}}"
        )

    def _toggle_theme(self):
        global _ACTIVE
        # Flip between dark and light
        if _ACTIVE.get("bg") == _DARK["bg"]:
            _ACTIVE = dict(_LIGHT)
            new_style = LIGHT_STYLE
            is_dark = False
        else:
            _ACTIVE = dict(_DARK)
            new_style = DARK_STYLE
            is_dark = True

        QApplication.instance().setStyleSheet(new_style)
        QSettings("vol3gui", "prefs").setValue("dark_mode", is_dark)

        self._apply_header_colours()

        # Update file chip colour if image loaded
        if self._dump_path:
            c = _c()
            self.hdr_file_chip.setStyleSheet(
                f"color:{c['success']};font-size:11px;background:transparent;")

        # Restyle all open result tabs
        for i in range(self.results_tabs.count()):
            tab = self.results_tabs.widget(i)
            if isinstance(tab, ResultsTab):
                tab.restyle()

        # Restyle browser category items
        self.browser.restyle()

        # Restyle coverage matrix if open
        if self._coverage_win and self._coverage_win.isVisible():
            self._coverage_win.restyle()

    def _build_menu(self):
        mb = self.menuBar()

        # ── File ──────────────────────────────────────────────────────────
        fm = mb.addMenu("&File")
        self._act(fm, "Open Image…", self._open_file, "Ctrl+O")
        fm.addSeparator()
        self._act(fm, "Open Case…", self._open_case, "Ctrl+Shift+O")
        self._act(fm, "Save Case…", self._save_case, "Ctrl+S")
        fm.addSeparator()
        self._act(fm, "Exit", self.close, "Ctrl+Q")

        # ── Symbols ───────────────────────────────────────────────────────
        sm = mb.addMenu("&Symbols")
        sm.setToolTipsVisible(True)

        linux_m = sm.addMenu("Linux Symbol Tables")
        self._act(linux_m, "Add Symbol Table File(s)…",
                  lambda: self._add_symbol_files("linux"))
        self._act(linux_m, "Add Symbol Table Directory…",
                  lambda: self._add_symbol_dir("linux"))
        linux_m.addSeparator()
        self._linux_list_action = self._act(linux_m, "Loaded: (none)",
                                            lambda: self._show_symbol_list("linux"))
        self._linux_list_action.setEnabled(False)
        self._act(linux_m, "Clear Linux Symbols",
                  lambda: self._clear_symbols("linux"))

        mac_m = sm.addMenu("macOS Symbol Tables")
        self._act(mac_m, "Add Symbol Table File(s)…",
                  lambda: self._add_symbol_files("mac"))
        self._act(mac_m, "Add Symbol Table Directory…",
                  lambda: self._add_symbol_dir("mac"))
        mac_m.addSeparator()
        self._mac_list_action = self._act(mac_m, "Loaded: (none)",
                                          lambda: self._show_symbol_list("mac"))
        self._mac_list_action.setEnabled(False)
        self._act(mac_m, "Clear macOS Symbols",
                  lambda: self._clear_symbols("mac"))

        sm.addSeparator()
        self._act(sm, "Clear All Custom Symbols", self._clear_all_symbols)

        # ── Results ───────────────────────────────────────────────────────
        rm = mb.addMenu("&Results")
        self._act(rm, "Close All Tabs", self._clear_tabs)
        self._act(rm, "Clear Log", self.log_panel._clear)

        # ── Plugins ───────────────────────────────────────────────────────
        pm = mb.addMenu("&Plugins")
        self._act(pm, "Refresh Plugin List", self._discover_plugins, "F5")

        # ── Tools ─────────────────────────────────────────────────────────
        tm = mb.addMenu("&Tools")
        self._act(tm, "Open Volshell…", self._open_volshell, "Ctrl+Shift+S")
        self._act(tm, "MITRE Coverage Matrix…",
                  self._open_coverage_matrix, "Ctrl+Shift+M")

        # ── Help ──────────────────────────────────────────────────────────
        hm = mb.addMenu("&Help")
        self._act(hm, "About", self._about)

    def _act(self, menu, label, slot, shortcut=None):
        a = QAction(label, self)
        if shortcut: a.setShortcut(shortcut)
        a.triggered.connect(slot)
        menu.addAction(a)
        return a

    def _build_statusbar(self):
        sb = self.statusBar()
        self._sb_main  = QLabel("Ready")
        self._sb_right = QLabel("")
        self._sb_right.setObjectName("runLabel")
        self._sb_stop_btn = QPushButton("■  Stop")
        self._sb_stop_btn.setFixedHeight(20)
        self._sb_stop_btn.setVisible(False)
        self._sb_stop_btn.clicked.connect(self._stop_runner)
        c = _c()
        self._sb_stop_btn.setStyleSheet(
            f"QPushButton{{background:transparent;color:{c['error']};"
            f"border:1px solid {c['error']};border-radius:3px;"
            f"font-size:11px;padding:0 8px;}}"
            f"QPushButton:hover{{background:{c['error']};color:{c['text_hi']};}}"
        )
        sb.addWidget(self._sb_main)
        sb.addPermanentWidget(self._sb_right)
        sb.addPermanentWidget(self._sb_stop_btn)

    # ── discovery ────────────────────────────────────────────────────────────

    def _discover_plugins(self):
        if not VOL3_OK:
            self.log_panel.log(f"volatility3 import failed: {VOL3_ERROR}", "error")
            return
        self.log_panel.log("Discovering plugins…", "debug")
        self._disc = PluginDiscoveryThread(self)
        self._disc.plugins_ready.connect(self._on_plugins_ready)
        self._disc.error.connect(lambda e: self.log_panel.log(e, "error"))
        self._disc.start()

    def _on_plugins_ready(self, cats: dict):
        total = sum(len(v) for v in cats.values())
        self._plugin_cats = cats
        self.browser.populate(cats)
        self.log_panel.log(f"{total} plugins loaded.", "success")
        self._sb_main.setText(f"{total} plugins")
        # Refresh coverage window if already open
        if self._coverage_win and self._coverage_win.isVisible():
            self._coverage_win.close()
            self._coverage_win = None

    # ── file ─────────────────────────────────────────────────────────────────

    def _open_file(self):
        settings = QSettings("vol3gui", "prefs")
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Memory Image", settings.value("last_dir", ""),
            "Memory Images (*.dmp *.mem *.vmem *.raw *.img *.bin *.lime *.dd *.E01 *.e01);;"
            "All Files (*)")
        if path:
            self.drop_zone._load(path)

    # ── case persistence ───────────────────────────────────────────────────────

    def _result_tabs(self) -> List["ResultsTab"]:
        return [self.results_tabs.widget(i)
                for i in range(self.results_tabs.count())
                if isinstance(self.results_tabs.widget(i), ResultsTab)]

    def _save_case(self):
        tabs = self._result_tabs()
        if not self._dump_path and not tabs:
            QMessageBox.information(
                self, "Nothing to Save",
                "Load an image and run at least one plugin first.")
            return
        default = ""
        if self._dump_path:
            default = str(Path(self._dump_path).with_suffix("")) + "_case.json"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Case", default, "Remora Case (*.json)")
        if not path:
            return
        case = {
            "remora_case_version": 1,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "image": {
                "path":   self._dump_path,
                "size":   self._image_meta.get("size"),
                "sha256": self._image_meta.get("sha256"),
                "md5":    self._image_meta.get("md5"),
            },
            "symbol_paths": self._symbol_paths,
            "results": [
                {
                    "plugin":    t.plugin_name,
                    "dump_path": t.dump_path,
                    "columns":   t.columns,
                    "rows":      t.rows,
                    "timestamp": t._ts,
                } for t in tabs
            ],
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(case, f, indent=2, default=str)
        except OSError as exc:
            QMessageBox.warning(self, "Save Failed", str(exc))
            return
        self.log_panel.log(
            f"Case saved: {path}  ({len(tabs)} result(s))", "success")

    def _open_case(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Case",
            QSettings("vol3gui", "prefs").value("last_dir", ""),
            "Remora Case (*.json);;All Files (*)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                case = json.load(f)
        except (OSError, json.JSONDecodeError) as exc:
            QMessageBox.warning(self, "Open Failed",
                                f"Could not read case file:\n{exc}")
            return
        if not isinstance(case, dict) or "results" not in case:
            QMessageBox.warning(self, "Invalid Case",
                                "This file is not a Remora case.")
            return

        img = case.get("image") or {}
        self._image_meta = {
            "path":   img.get("path"),
            "size":   img.get("size"),
            "sha256": img.get("sha256"),
            "md5":    img.get("md5"),
        }
        self._dump_path = img.get("path")

        sp = case.get("symbol_paths") or {}
        self._symbol_paths = {
            "linux": list(sp.get("linux", [])),
            "mac":   list(sp.get("mac", [])),
        }
        self._refresh_symbol_labels()

        # Header chip reflects the restored image + its recorded hash
        if self._dump_path:
            fname = os.path.basename(self._dump_path)
            size  = _fmt_size(img.get("size") or 0)
            sha   = img.get("sha256") or ""
            c = _c()
            chip = f"{fname}  ·  {size}"
            if sha:
                chip += f"  ·  {sha[:12]}…"
            self.hdr_file_chip.setText(chip)
            self.hdr_file_chip.setStyleSheet(
                f"color:{c['success']};font-size:11px;background:transparent;")
            if sha:
                self.hdr_file_chip.setToolTip(
                    f"SHA-256: {sha}\nMD5: {img.get('md5') or ''}")

        # Only allow new runs when the original image is actually on disk
        exists = bool(self._dump_path) and os.path.isfile(self._dump_path)
        self.config_panel.enable_run(exists)
        if self._dump_path and not exists:
            self.log_panel.log(
                f"Image not found on disk: {self._dump_path} — "
                f"results loaded read-only.", "warning")

        # Rebuild result tabs from the saved data (no plugins are re-run)
        self._clear_tabs()
        n = 0
        for r in case.get("results", []):
            if self.results_tabs.count() == 1 and self.results_tabs.tabText(0) == "·":
                self.results_tabs.removeTab(0)
            tab = ResultsTab(
                r.get("plugin", ""), r.get("dump_path", ""),
                r.get("columns", []) or [], r.get("rows", []) or [],
                image_meta=self._image_meta, ts=r.get("timestamp"))
            self.results_tabs.addTab(tab, r.get("plugin", "").split(".")[-1])
            n += 1
        if n:
            self.results_tabs.setCurrentIndex(self.results_tabs.count() - 1)
        self.log_panel.log(
            f"Case loaded: {path}  ({n} result(s))", "success")
        self._sb_main.setText(
            f"Case: {os.path.basename(path)} — {n} result(s)")

    def _on_dump_loaded(self, path: str):
        self._dump_path = path
        fname = os.path.basename(path)
        raw_size = os.path.getsize(path)
        size  = _fmt_size(raw_size)
        c = _c()
        self.hdr_file_chip.setText(f"{fname}  ·  {size}")
        self.hdr_file_chip.setStyleSheet(
            f"color:{c['success']};font-size:11px;background:transparent;")
        self.log_panel.set_log_file(path)
        self.log_panel.log(f"Loaded: {path}", "success")
        self._sb_main.setText(f"{fname}  ({size})")
        self.config_panel.enable_run(True)
        # Begin chain-of-custody hashing of the image
        self._start_hashing(path, raw_size)

    # ── evidence integrity (chain of custody) ─────────────────────────────────

    def _start_hashing(self, path: str, size: int):
        if self._hash_thread and self._hash_thread.isRunning():
            self._hash_thread.abort()
            self._hash_thread.wait(2000)
        self._image_meta = {"path": path, "size": size,
                            "md5": None, "sha256": None}
        self.log_panel.log("Hashing image for chain of custody…", "debug")
        self._sb_right.setText("Hashing 0%…  ")
        self._hash_thread = ImageHashThread(path, self)
        self._hash_thread.progress.connect(self._on_hash_progress)
        self._hash_thread.done.connect(self._on_hash_done)
        self._hash_thread.start()

    def _on_hash_progress(self, pct: int):
        # Don't clobber the run indicator while a plugin is executing
        if not (self._runner and self._runner.isRunning()):
            self._sb_right.setText(f"Hashing {pct}%…  ")

    def _on_hash_done(self, result: dict):
        if not (self._runner and self._runner.isRunning()):
            self._sb_right.setText("")
        if result.get("algo_error"):
            self.log_panel.log(
                f"Image hashing failed: {result['algo_error']}", "warning")
            return
        self._image_meta.update(result)
        sha = result.get("sha256", "")
        md5 = result.get("md5", "")
        # Audit lines — written verbatim into the persistent log file
        self.log_panel.log(f"Image SHA-256: {sha}", "info")
        self.log_panel.log(f"Image MD5    : {md5}", "info")
        # Surface a short digest in the header chip tooltip + text
        if self._dump_path:
            fname = os.path.basename(self._dump_path)
            size  = _fmt_size(self._image_meta.get("size", 0))
            self.hdr_file_chip.setText(f"{fname}  ·  {size}  ·  {sha[:12]}…")
            self.hdr_file_chip.setToolTip(
                f"SHA-256: {sha}\nMD5: {md5}\nSize: {self._image_meta.get('size', 0):,} bytes")

    # ── plugin selection / run ────────────────────────────────────────────────

    def _on_plugin_selected(self, name: str, cls):
        self.config_panel.load_plugin(name, cls)
        if self._dump_path:
            self.config_panel.enable_run(True)

    def _quick_run(self, name: str, cls):
        self.config_panel.load_plugin(name, cls)
        if not self._dump_path:
            QMessageBox.warning(self, "No Image",
                "Load a memory image first (drag & drop or File > Open).")
            return
        self._on_run(name, cls, {}, "")

    def _on_run(self, plugin_name: str, cls, args: dict, out_dir: str):
        if not self._dump_path:
            QMessageBox.warning(self, "No Image", "Load a memory image first.")
            return
        if self._runner and self._runner.isRunning():
            if QMessageBox.question(
                self, "Plugin Running",
                "Stop the current plugin and run the new one?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
                return
            # A manual run cancels any queued batch
            self._clear_batch()
            self._runner.abort()
            self._runner.wait(4000)
        self._start_runner(plugin_name, args, out_dir)

    def _start_runner(self, plugin_name: str, args: dict, out_dir: str):
        self.log_panel.log(f"→ {plugin_name}", "info")
        self._set_busy(True, plugin_name)
        self._runner = PluginRunnerThread(
            plugin_name, self._dump_path, args, out_dir,
            symbol_paths=self._symbol_paths, parent=self)
        self._runner.log_line.connect(self.log_panel.log)
        self._runner.result_data.connect(self._on_result)
        self._runner.finished.connect(self._on_done)
        self._runner.start()

    # ── batch / triage execution ──────────────────────────────────────────────

    def _on_profile_applied(self, profile: str, matched: int):
        if matched:
            self.log_panel.log(
                f"Triage profile '{profile}': checked {matched} plugin(s). "
                f"Click Run Checked to execute.", "info")
        else:
            self.log_panel.log(
                f"Triage profile '{profile}': no matching plugins found "
                f"(wait for plugin discovery, or check the target OS).",
                "warning")

    def _on_batch_run(self, plugins: list):
        if not self._dump_path:
            QMessageBox.warning(self, "No Image", "Load a memory image first.")
            return
        if self._runner and self._runner.isRunning():
            QMessageBox.warning(
                self, "Busy", "Wait for the current plugin to finish first.")
            return
        if not plugins:
            return
        self._batch_queue = list(plugins)
        self._batch_total = len(plugins)
        self.log_panel.log(
            f"Batch: queued {self._batch_total} plugin(s) with defaults.", "info")
        self._advance_batch()

    def _advance_batch(self):
        if not self._batch_queue:
            if self._batch_total:
                self.log_panel.log(
                    f"Batch complete — {self._batch_total} plugin(s) run.", "success")
                self._batch_total = 0
            return
        name, _cls = self._batch_queue.pop(0)
        done_n = self._batch_total - len(self._batch_queue)
        self.log_panel.log(f"Batch [{done_n}/{self._batch_total}]", "debug")
        self._start_runner(name, {}, "")

    def _clear_batch(self):
        if self._batch_queue or self._batch_total:
            self.log_panel.log("Batch cancelled.", "warning")
        self._batch_queue = []
        self._batch_total = 0

    def _on_result(self, plugin, dump, cols, rows):
        if self.results_tabs.count() == 1 and self.results_tabs.tabText(0) == "·":
            self.results_tabs.removeTab(0)
        self._tab_n += 1
        tab = ResultsTab(plugin, dump, cols, rows, image_meta=self._image_meta)
        idx = self.results_tabs.addTab(tab, plugin.split(".")[-1])
        self.results_tabs.setCurrentIndex(idx)
        self.log_panel.log(f"← {plugin.split('.')[-1]}  {len(rows):,} rows", "success")

    def _on_done(self, ok: bool):
        self._set_busy(False, "")
        if not ok:
            self.log_panel.log("Plugin finished with errors.", "warning")
        # Drive the batch queue forward (no-op for single runs)
        self._advance_batch()

    # ── helpers ───────────────────────────────────────────────────────────────

    def _set_busy(self, busy: bool, plugin: str):
        self.progress.setVisible(busy)
        self._sb_stop_btn.setVisible(busy)
        if busy:
            self.progress.setRange(0, 0)
            short = plugin.split(".")[-1] if plugin else "plugin"
            self.hdr_run_lbl.setText(f"  ▶  {short}")
            self._sb_right.setText(f"Running {short}…  ")
        else:
            self.progress.setRange(0, 1)
            self.progress.setValue(1)
            self.hdr_run_lbl.setText("")
            self._sb_right.setText("")

    def _stop_runner(self):
        if self._runner and self._runner.isRunning():
            self._clear_batch()
            self._runner.abort()
            self.log_panel.log("Plugin stopped by user.", "warning")
            self._set_busy(False, "")

    def _show_welcome(self):
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        c = _c()
        accent   = c["accent"]
        t_muted  = c["text_muted"]
        lbl = QLabel(
            f"<div style='text-align:center;font-family:monospace'>"
            f"<div style='font-size:11px;font-weight:700;letter-spacing:3px;"
            f"color:{accent};margin-bottom:16px'>[ NO RESULTS ]</div>"
            f"<div style='font-size:11px;line-height:2;color:{t_muted}'>"
            f"01 &nbsp; load a memory image<br>"
            f"02 &nbsp; select a plugin<br>"
            f"03 &nbsp; run"
            f"</div></div>")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet(f"color:{c['text_dim']};")
        lay.addWidget(lbl)
        self.results_tabs.addTab(w, "·")
        self.results_tabs.tabBar().setTabButton(0, self.results_tabs.tabBar().ButtonPosition.RightSide, None)

    def _close_tab(self, idx: int):
        tab = self.results_tabs.widget(idx)
        if isinstance(tab, VolshellWidget):
            tab._kill()
            self._volshell_tab_idx = -1
        self.results_tabs.removeTab(idx)
        if self.results_tabs.count() == 0:
            self._show_welcome()

    def _on_tab_close_track(self, idx: int):
        # Sync _volshell_tab_idx when tabs shift after close
        if self._volshell_tab_idx > idx:
            self._volshell_tab_idx -= 1
        elif self._volshell_tab_idx == idx:
            self._volshell_tab_idx = -1

    def _clear_tabs(self):
        for i in range(self.results_tabs.count()):
            tab = self.results_tabs.widget(i)
            if isinstance(tab, VolshellWidget):
                tab._kill()
        self.results_tabs.clear()
        self._volshell_tab_idx = -1
        self._show_welcome()

    def _about(self):
        QMessageBox.about(self, "Remora",
            "<h3>Remora</h3>"
            "<p>Graphical frontend for the Volatility3 memory forensics framework.</p>"
            "<p><b>Exports:</b> CSV · TSV · JSON · TXT · HTML · PDF · XLSX</p>"
            "<p>Drag &amp; drop a memory image to begin.</p>")

    # ── Symbol table management ───────────────────────────────────────────────

    def _add_symbol_files(self, os_type: str):
        paths, _ = QFileDialog.getOpenFileNames(
            self, f"Select {os_type.title()} Symbol Table(s)",
            QSettings("vol3gui", "prefs").value("last_sym_dir", ""),
            "Symbol Tables (*.json *.json.gz *.isf *.isf.gz);;All Files (*)")
        if not paths:
            return
        QSettings("vol3gui", "prefs").setValue(
            "last_sym_dir", str(Path(paths[0]).parent))
        added = 0
        for p in paths:
            if p not in self._symbol_paths[os_type]:
                self._symbol_paths[os_type].append(p)
                added += 1
        self._refresh_symbol_labels()
        self.log_panel.log(
            f"Added {added} {os_type} symbol table(s). "
            f"Total: {len(self._symbol_paths[os_type])}", "success")

    def _add_symbol_dir(self, os_type: str):
        d = QFileDialog.getExistingDirectory(
            self, f"Select {os_type.title()} Symbol Table Directory",
            QSettings("vol3gui", "prefs").value("last_sym_dir", ""))
        if not d:
            return
        QSettings("vol3gui", "prefs").setValue("last_sym_dir", d)
        if d not in self._symbol_paths[os_type]:
            self._symbol_paths[os_type].append(d)
        self._refresh_symbol_labels()
        self.log_panel.log(
            f"Added {os_type} symbol directory: {d}", "success")

    def _clear_symbols(self, os_type: str):
        self._symbol_paths[os_type].clear()
        self._refresh_symbol_labels()
        self.log_panel.log(f"Cleared {os_type} symbol paths.", "info")

    def _clear_all_symbols(self):
        self._symbol_paths = {"linux": [], "mac": []}
        self._refresh_symbol_labels()
        self.log_panel.log("Cleared all custom symbol paths.", "info")

    def _show_symbol_list(self, os_type: str):
        paths = self._symbol_paths[os_type]
        if not paths:
            QMessageBox.information(self, f"{os_type.title()} Symbols",
                                    "No custom symbol paths loaded.")
            return
        msg = "\n".join(f"  {i+1}.  {p}" for i, p in enumerate(paths))
        QMessageBox.information(
            self, f"{os_type.title()} Symbol Paths ({len(paths)})", msg)

    def _refresh_symbol_labels(self):
        for os_type, action_attr in (
                ("linux", "_linux_list_action"),
                ("mac", "_mac_list_action")):
            paths = self._symbol_paths[os_type]
            action = getattr(self, action_attr)
            if paths:
                action.setText(f"Loaded: {len(paths)} path(s) — view…")
                action.setEnabled(True)
            else:
                action.setText("Loaded: (none)")
                action.setEnabled(False)

    # ── Volshell ──────────────────────────────────────────────────────────────

    def _open_volshell(self):
        if not self._dump_path:
            QMessageBox.warning(self, "No Image",
                "Load a memory image first (drag & drop or File > Open).")
            return

        # If a volshell tab already exists, just switch to it
        if self._volshell_tab_idx >= 0:
            tab = self.results_tabs.widget(self._volshell_tab_idx)
            if isinstance(tab, VolshellWidget):
                self.results_tabs.setCurrentIndex(self._volshell_tab_idx)
                return
            # tab was closed — reset
            self._volshell_tab_idx = -1

        # Remove welcome tab if present
        if self.results_tabs.count() == 1 and self.results_tabs.tabText(0) == "·":
            self.results_tabs.removeTab(0)

        vw = VolshellWidget(self._dump_path)
        idx = self.results_tabs.addTab(vw, "Volshell")
        self.results_tabs.setCurrentIndex(idx)
        self._volshell_tab_idx = idx

    # ── MITRE Coverage Matrix ─────────────────────────────────────────────────

    def _open_coverage_matrix(self):
        if not self._plugin_cats:
            QMessageBox.information(
                self, "No Plugins",
                "Wait for plugin discovery to complete (F5 to refresh).")
            return
        if self._coverage_win and self._coverage_win.isVisible():
            self._coverage_win.raise_()
            self._coverage_win.activateWindow()
            return
        self._coverage_win = MitreCoverageDialog(self._plugin_cats, parent=None)
        QApplication.instance().setStyleSheet(
            QApplication.instance().styleSheet())  # ensure theme propagates
        self._coverage_win.show()

    def closeEvent(self, event):
        if self._runner and self._runner.isRunning():
            self._runner.abort()
            self._runner.wait(3000)
        if self._hash_thread and self._hash_thread.isRunning():
            self._hash_thread.abort()
            self._hash_thread.wait(2000)
        if self._coverage_win:
            self._coverage_win.close()
        event.accept()


# ===========================================================================
# Utilities
# ===========================================================================

def _fmt_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024: return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} PB"

def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _esc(text: str) -> str:
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def _done(parent: QWidget, path: str):
    QMessageBox.information(parent, "Exported", f"Saved to:\n{path}")

def _ghost_btn(label: str) -> QPushButton:
    btn = QPushButton(label)
    btn.setProperty("class", "ghost")
    # Inline override ensures ghost style even before stylesheet propagates
    btn.setStyleSheet(
        "QPushButton[class='ghost']{background:transparent;border:1px solid palette(mid);"
        "border-radius:2px;padding:3px 10px;font-size:11px;min-height:24px}"
    )
    return btn


# ===========================================================================
# Entry point
# ===========================================================================

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Remora")
    app.setOrganizationName("vol3gui")
    app.setFont(QFont("Inter", 11))

    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
